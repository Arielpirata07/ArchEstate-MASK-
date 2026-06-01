# app.py - Aplicación principal de Flask para ArchEstate

import csv
import io
import os
import re
import secrets
import sqlite3
import threading
import time

from datetime import datetime, timedelta
from functools import wraps
from io import StringIO

import openpyxl
import phonenumbers
from phonenumbers import PhoneNumberType
import pytz

from flask import Flask, render_template, jsonify, request, session, redirect, url_for, Response, send_file, flash, send_from_directory
from fpdf import FPDF
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

import config
import utils
from utils import allowed_file, convert_to_argentina_time
import decorators
import rate_limit
import validators
import models


app = Flask(__name__)
app.secret_key = config.SECRET_KEY
app.config['UPLOAD_FOLDER'] = os.path.join(app.root_path, config.UPLOAD_FOLDER)
app.config['AVATAR_FOLDER'] = os.path.join(app.root_path, config.AVATAR_FOLDER)
app.config['PERMANENT_SESSION_LIFETIME'] = config.PERMANENT_SESSION_LIFETIME
app.jinja_env.autoescape = True


class FilterOptionsCache:
    """Caché simple en memoria para opciones de filtros de leads"""
    def __init__(self, ttl_seconds=300):
        self._cache = {}
        self._timestamps = {}
        self._ttl = ttl_seconds
        self._lock = threading.Lock()

    def get(self, key):
        with self._lock:
            if key in self._cache:
                if time.time() - self._timestamps[key] < self._ttl:
                    return self._cache[key]
                else:
                    del self._cache[key]
                    del self._timestamps[key]
        return None

    def set(self, key, value):
        with self._lock:
            self._cache[key] = value
            self._timestamps[key] = time.time()

    def invalidate(self):
        with self._lock:
            self._cache.clear()
            self._timestamps.clear()


filter_cache = FilterOptionsCache(ttl_seconds=300)


@app.after_request
def security_headers(response):
    """Agrega headers de seguridad HTTP a todas las respuestas"""
    if not request.path.startswith('/static/'):
        response.headers['X-Content-Type-Options'] = 'nosniff'
        response.headers['X-Frame-Options'] = 'DENY'
        response.headers['X-XSS-Protection'] = '1; mode=block'
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
        response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
        response.headers['Permissions-Policy'] = 'camera=(), microphone=(), geolocation=()'
    return rate_limit.add_rate_limit_headers(response)


@app.before_request
def assign_request_id():
    """Asigna un request_id por request para trazabilidad en errores."""
    import uuid
    from flask import g
    g.request_id = uuid.uuid4().hex[:12]


@app.before_request
def restore_session_from_remember_cookie():
    """
    Si no hay sesión activa pero la cookie remember_token es válida y vigente,
    restaura la sesión del usuario y la marca como permanente.
    Rutas exentas: /login, /register, /logout, /api/auth/* y assets.
    """
    from flask import g

    if session.get('user_id'):
        return None

    path = request.path or ''
    exempt_prefixes = ('/static/', '/login', '/register', '/logout', '/api/auth/', '/sitemap.xml', '/robots.txt')
    if any(path == p or path.startswith(p) for p in exempt_prefixes):
        return None

    raw_cookie = request.cookies.get(config.REMEMBER_COOKIE_NAME)
    if not raw_cookie or ':' not in raw_cookie:
        return None

    selector, _, validator = raw_cookie.partition(':')
    if not selector or not validator:
        return None

    user_id = utils.validate_remember_token(selector, validator)
    if not user_id:
        return None

    conn = None
    try:
        conn = get_db_connection()
        user = conn.execute(
            'SELECT id, username, email, role, is_active FROM users WHERE id = ?',
            (user_id,)
        ).fetchone()
        if not user or not user['is_active']:
            utils.revoke_remember_token(selector)
            return None

        session.permanent = True
        session['user_id'] = user['id']
        session['username'] = user['username']
        session['email'] = user['email']
        session['role'] = user['role']
        g.restored_from_remember = True
        utils.log_event(
            user_id=user['id'], event='remember_session_restored',
            props={'selector_prefix': selector[:8]}
        )
    except Exception as e:
        print(f"Error al restaurar sesión desde remember token: {e}")
    finally:
        if conn:
            conn.close()
    return None


@app.context_processor
def inject_request_id():
    """Hace request_id disponible en todas las templates."""
    from flask import g
    return dict(request_id=getattr(g, 'request_id', None))


def _err_response(status, message, code=None):
    """Helper para construir respuestas de error con request_id."""
    from flask import g, jsonify
    body = {
        "error": message,
        "code": code or str(status),
        "request_id": getattr(g, 'request_id', None),
    }
    return jsonify(body), status


@app.errorhandler(400)
def _h_400(e): return _err_response(400, "Solicitud malformada.", "BAD_REQUEST")

@app.errorhandler(404)
def _h_404(e): return _err_response(404, "Recurso no encontrado.", "NOT_FOUND")

@app.errorhandler(409)
def _h_409(e): return _err_response(409, "Conflicto con el estado actual.", "CONFLICT")

@app.errorhandler(410)
def _h_410(e): return _err_response(410, "Recurso expirado.", "GONE")

@app.errorhandler(429)
def _h_429(e): return _err_response(429, "Demasiadas solicitudes.", "RATE_LIMITED")

@app.errorhandler(500)
def _h_500(e):
    from flask import g
    rid = getattr(g, 'request_id', None)
    print(f"[500] request_id={rid}: {e}")
    return _err_response(500, "Error interno del servidor.", "INTERNAL")


def get_db_connection():
    return models.get_db_connection()


def init_db():
    """Inicializa la base de datos de usuarios, leads, profesionales y auditoría si no existe"""
    with app.app_context():
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Tabla de Usuarios
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL UNIQUE,
                email TEXT NOT NULL DEFAULT '',
                hash TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'client',
                doc_path TEXT DEFAULT '',
                is_active INTEGER NOT NULL DEFAULT 1
            )
        ''')
        
        # Migraciones de columnas faltantes
        cursor.execute('PRAGMA table_info(users)')
        user_columns = [row[1] for row in cursor.fetchall()]
        if 'email' not in user_columns:
            cursor.execute("ALTER TABLE users ADD COLUMN email TEXT NOT NULL DEFAULT ''")
        if 'phone' not in user_columns:
            cursor.execute("ALTER TABLE users ADD COLUMN phone TEXT NOT NULL DEFAULT ''")
        if 'phone_format_valid' not in user_columns:
            cursor.execute("ALTER TABLE users ADD COLUMN phone_format_valid INTEGER DEFAULT 0")
        if 'is_active' not in user_columns:
            cursor.execute("ALTER TABLE users ADD COLUMN is_active INTEGER NOT NULL DEFAULT 1")
        if 'phone_verified' not in user_columns:
            cursor.execute("ALTER TABLE users ADD COLUMN phone_verified INTEGER DEFAULT 0")
        if 'verification_code' not in user_columns:
            cursor.execute("ALTER TABLE users ADD COLUMN verification_code TEXT DEFAULT ''")
        if 'verification_expires' not in user_columns:
            cursor.execute("ALTER TABLE users ADD COLUMN verification_expires DATETIME")
        if 'phone_e164' not in user_columns:
            cursor.execute("ALTER TABLE users ADD COLUMN phone_e164 TEXT DEFAULT ''")
        if 'phone_number_type' not in user_columns:
            cursor.execute("ALTER TABLE users ADD COLUMN phone_number_type TEXT DEFAULT ''")

        # Tabla de Leads
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS leads (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                type TEXT NOT NULL,
                property_type TEXT NOT NULL DEFAULT 'departamento',
                zone TEXT NOT NULL,
                budget TEXT NOT NULL,
                currency TEXT NOT NULL DEFAULT 'ARG',
                phone TEXT NOT NULL,
                email TEXT NOT NULL,
                floor_block TEXT DEFAULT '',
                usable_m2 INTEGER DEFAULT 0,
                elevator TEXT DEFAULT '',
                land_area INTEGER DEFAULT 0,
                built_area INTEGER DEFAULT 0,
                pool TEXT DEFAULT '',
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Actualizar esquemas antiguos sin las nuevas columnas
        cursor.execute('PRAGMA table_info(leads)')
        existing_columns = [row[1] for row in cursor.fetchall()]
        schema_updates = [
            ('property_type', "TEXT NOT NULL DEFAULT 'departamento'"),
            ('floor_block', "TEXT DEFAULT ''"),
            ('usable_m2', "INTEGER DEFAULT 0"),
            ('elevator', "TEXT DEFAULT ''"),
            ('land_area', "INTEGER DEFAULT 0"),
            ('built_area', "INTEGER DEFAULT 0"),
            ('pool', "TEXT DEFAULT ''"),
            ('architectural_style', "TEXT DEFAULT ''"),
            ('bedrooms', "INTEGER DEFAULT 0"),
            ('bathrooms', "INTEGER DEFAULT 0"),
            ('total_area', "INTEGER DEFAULT 0"),
            ('amenities', "TEXT DEFAULT ''"),
            ('ambientes', "INTEGER DEFAULT 0"),
            ('parking', "TEXT DEFAULT ''"),
            ('orientation', "TEXT DEFAULT ''"),
            ('property_condition', "TEXT DEFAULT ''"),
            ('property_age', "TEXT DEFAULT ''"),
        ]
        for column, column_type in schema_updates:
            if column not in existing_columns:
                cursor.execute(f"ALTER TABLE leads ADD COLUMN {column} {column_type}")

        # Migración: provincia
        if 'province' not in existing_columns:
            cursor.execute("ALTER TABLE leads ADD COLUMN province TEXT DEFAULT ''")

        # Migración: phone_format_valid (Fase 1 — validación de formato)
        if 'phone_format_valid' not in existing_columns:
            cursor.execute("ALTER TABLE leads ADD COLUMN phone_format_valid INTEGER DEFAULT 0")

        # Migración: agregar user_id a leads si no existe
        cursor.execute('PRAGMA table_info(leads)')
        lead_columns = [row[1] for row in cursor.fetchall()]
        if 'user_id' not in lead_columns:
            cursor.execute('ALTER TABLE leads ADD COLUMN user_id INTEGER DEFAULT NULL')
            cursor.execute('''
                UPDATE leads SET user_id = (
                    SELECT u.id FROM users u WHERE u.email = leads.email
                ) WHERE user_id IS NULL AND leads.email IN (SELECT email FROM users)
            ''')

        # Tabla de Perfiles de Usuario
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS user_profiles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL UNIQUE REFERENCES users(id),
                first_name TEXT DEFAULT '',
                last_name TEXT DEFAULT '',
                bio TEXT DEFAULT '',
                title TEXT DEFAULT '',
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Tabla de Versiones de Leads
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS lead_versions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                lead_id INTEGER NOT NULL REFERENCES leads(id),
                version INTEGER NOT NULL,
                data_snapshot TEXT NOT NULL,
                created_by INTEGER REFERENCES users(id),
                change_summary TEXT DEFAULT '',
                edited_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Tabla de Profesionales
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS professionals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER DEFAULT NULL,
                name TEXT NOT NULL,
                license TEXT NOT NULL UNIQUE,
                specialty TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'pending',
                FOREIGN KEY (user_id) REFERENCES users(id)
            )
        ''')

        # Migración: agregar user_id si no existe
        cursor.execute('PRAGMA table_info(professionals)')
        pro_columns = [row[1] for row in cursor.fetchall()]
        if 'user_id' not in pro_columns:
            cursor.execute('ALTER TABLE professionals ADD COLUMN user_id INTEGER DEFAULT NULL')
            # Poblar user_id para profesionales cuyo name coincide con username
            cursor.execute('''
                UPDATE professionals SET user_id = (
                    SELECT u.id FROM users u WHERE u.username = professionals.name
                ) WHERE user_id IS NULL
            ''')

        # Tabla de Auditoría
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS audit_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                action TEXT NOT NULL,
                target TEXT NOT NULL,
                admin TEXT NOT NULL
            )
        ''')

        # Tabla de Tracking de Leads por Profesional
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS lead_tracking (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                professional_id INTEGER NOT NULL,
                lead_id INTEGER NOT NULL,
                seen INTEGER NOT NULL DEFAULT 0,
                contacted INTEGER NOT NULL DEFAULT 0,
                seen_at DATETIME DEFAULT NULL,
                contacted_at DATETIME DEFAULT NULL,
                UNIQUE(professional_id, lead_id),
                FOREIGN KEY (professional_id) REFERENCES users(id),
                FOREIGN KEY (lead_id) REFERENCES leads(id)
            )
        ''')

        # Tabla de Reportes de Leads por Profesionales
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS lead_reports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                lead_id INTEGER NOT NULL,
                reported_by INTEGER NOT NULL,
                reason TEXT NOT NULL DEFAULT 'telefono_inexistente',
                notes TEXT DEFAULT '',
                status TEXT NOT NULL DEFAULT 'pending',
                reviewed_by TEXT DEFAULT NULL,
                reviewed_at DATETIME DEFAULT NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (lead_id) REFERENCES leads(id),
                FOREIGN KEY (reported_by) REFERENCES users(id)
            )
        ''')

        # Tabla de Perfiles Profesionales Extendidos
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS professional_profiles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL UNIQUE REFERENCES users(id),
                photo_path TEXT DEFAULT '',
                bio_pro TEXT DEFAULT '',
                experience_years INTEGER DEFAULT 0,
                services_offered TEXT DEFAULT '[]',
                portfolio TEXT DEFAULT '[]',
                availability TEXT DEFAULT '{}',
                social_links TEXT DEFAULT '{}',
                fee_range_min REAL DEFAULT 0,
                fee_range_max REAL DEFAULT 0,
                professional_address TEXT DEFAULT '',
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Tabla de Preferencias de Usuario
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS user_preferences (
                user_id INTEGER PRIMARY KEY REFERENCES users(id),
                theme TEXT NOT NULL DEFAULT 'light',
                language TEXT NOT NULL DEFAULT 'es',
                email_notifications INTEGER NOT NULL DEFAULT 1,
                sms_notifications INTEGER NOT NULL DEFAULT 1,
                lead_alerts INTEGER NOT NULL DEFAULT 1,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Tabla de Historial de Sesiones (auditoría, no enforcement)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS user_login_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL REFERENCES users(id),
                ip_address TEXT DEFAULT '',
                user_agent TEXT DEFAULT '',
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                last_active DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Tabla de consentimiento de canales (Ley 25.326 / AAIP)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS consent_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL REFERENCES users(id),
                channel TEXT NOT NULL,
                ip TEXT DEFAULT '',
                user_agent TEXT DEFAULT '',
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Tabla genérica de eventos (telemetría)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER REFERENCES users(id),
                lead_id INTEGER REFERENCES leads(id),
                event TEXT NOT NULL,
                props_json TEXT DEFAULT '',
                ip TEXT DEFAULT '',
                ts DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS remember_tokens (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                selector TEXT NOT NULL UNIQUE,
                validator_hash TEXT NOT NULL,
                expires_at DATETIME NOT NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                ip_address TEXT DEFAULT '',
                user_agent TEXT DEFAULT '',
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            )
        ''')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_remember_tokens_selector ON remember_tokens(selector)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_remember_tokens_expires ON remember_tokens(expires_at)')

        # Migración: preferred_channel en user_preferences
        cursor.execute('PRAGMA table_info(user_preferences)')
        up_prefs_cols = [r[1] for r in cursor.fetchall()]
        if 'preferred_channel' not in up_prefs_cols:
            cursor.execute("ALTER TABLE user_preferences ADD COLUMN preferred_channel TEXT NOT NULL DEFAULT 'auto'")

        # Backfill: derivar phone_e164 y phone_number_type para users existentes
        cursor.execute('''
            SELECT id, phone FROM users
            WHERE (phone_e164 IS NULL OR phone_e164 = '')
              AND phone IS NOT NULL AND phone != ''
        ''')
        pending = cursor.fetchall()
        for row in pending:
            e164 = utils.normalize_phone_to_e164(row['phone'])
            ntype = ''
            if e164 and utils.is_mobile_number(e164):
                parsed = utils._parse_phone(e164)
                if parsed is not None:
                    t = phonenumbers.number_type(parsed)
                    ntype = 'mobile' if t == PhoneNumberType.MOBILE else (
                        'fixed_or_mobile' if t == PhoneNumberType.FIXED_LINE_OR_MOBILE else 'fixed')
            cursor.execute(
                'UPDATE users SET phone_e164 = ?, phone_number_type = ? WHERE id = ?',
                (e164, ntype, row['id'])
            )
        if pending:
            print(f"[init_db] Backfill phone_e164 para {len(pending)} usuarios")

        # Migraciones adicionales en tablas existentes
        cursor.execute('PRAGMA table_info(user_profiles)')
        up_cols = [r[1] for r in cursor.fetchall()]
        if 'avatar_path' not in up_cols:
            cursor.execute("ALTER TABLE user_profiles ADD COLUMN avatar_path TEXT DEFAULT ''")

        cursor.execute('PRAGMA table_info(professionals)')
        pro_cols = [r[1] for r in cursor.fetchall()]
        if 'license_verified' not in pro_cols:
            cursor.execute("ALTER TABLE professionals ADD COLUMN license_verified INTEGER NOT NULL DEFAULT 0")

        cursor.execute('PRAGMA table_info(audit_log)')
        al_cols = [r[1] for r in cursor.fetchall()]
        if 'user_id' not in al_cols:
            cursor.execute("ALTER TABLE audit_log ADD COLUMN user_id INTEGER REFERENCES users(id)")

        # Crear directorios de uploads
        os.makedirs(config.AVATAR_FOLDER, exist_ok=True)
        os.makedirs(os.path.join('static', 'uploads', 'portfolio'), exist_ok=True)

        # Indices para optimizar consultas frecuentes
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_leads_user_id ON leads(user_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_leads_timestamp ON leads(timestamp)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_leads_zone ON leads(zone)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_leads_type ON leads(type)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_professionals_user_id ON professionals(user_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_professionals_name ON professionals(name)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_professionals_status ON professionals(status)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_lead_tracking_professional ON lead_tracking(professional_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_lead_reports_status ON lead_reports(status)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_audit_log_user ON audit_log(user_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_login_history_user ON user_login_history(user_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_events_user ON events(user_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_events_lead ON events(lead_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_events_event ON events(event)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_events_ts ON events(ts)')

        # CREAMOS EL ADMIN POR DEFECTO (Ahora con su rol)
        cursor.execute('SELECT COUNT(*) FROM users')
        if cursor.fetchone()[0] == 0:
            cursor.execute('INSERT INTO users (username, email, hash, role) VALUES (?, ?, ?, ?)', 
                          ('admin', 'admin@archestate.local', generate_password_hash('admin123'), 'admin'))
        conn.commit()
        conn.close()


# --- DECORADORES (from decorators.py) ---
login_required = decorators.login_required
admin_required = decorators.admin_required
professional_required = decorators.professional_required


# --- LÓGICA DE NEGOCIO (PYTHON) ---


def get_budget_stats_from_db():
    """Retorna estadísticas de presupuesto desde la base de datos"""
    conn = None
    try:
        conn = get_db_connection()
        total_leads = conn.execute('SELECT COUNT(*) FROM leads').fetchone()[0]
        leads_by_budget = conn.execute(
            'SELECT budget, COUNT(*) as count FROM leads GROUP BY budget ORDER BY count DESC'
        ).fetchall()
        leads_by_currency = conn.execute(
            'SELECT currency, COUNT(*) as count FROM leads GROUP BY currency'
        ).fetchall()
        return {
            'total_leads': total_leads,
            'by_budget': [{'label': r['budget'], 'value': r['count']} for r in leads_by_budget],
            'by_currency': [{'label': r['currency'], 'value': r['count']} for r in leads_by_currency],
        }
    finally:
        if conn:
            conn.close()

# --- RUTAS DE NAVEGACIÓN (VISTAS) ---

@app.route('/')
def index():
    return render_template('landing.html')


@app.route('/api/landing/stats', methods=['GET'])
def landing_stats():
    """Retorna estadisticas publicas para la landing page (sin auth)."""
    conn = None
    try:
        conn = get_db_connection()

        # Total de leads
        total_leads = conn.execute('SELECT COUNT(*) FROM leads').fetchone()[0]

        # Profesionales aprobados
        total_pros = conn.execute(
            "SELECT COUNT(*) FROM professionals WHERE status = 'approved'"
        ).fetchone()[0]

        # Zonas unicas cubiertas
        total_zones = conn.execute(
            'SELECT COUNT(DISTINCT zone) FROM leads WHERE zone != ""'
        ).fetchone()[0]

        # Leads del mes actual
        leads_this_month = conn.execute('''
            SELECT COUNT(*) FROM leads
            WHERE strftime('%Y-%m', timestamp) = strftime('%Y-%m', 'now')
        ''').fetchone()[0]

        return jsonify({
            'total_leads': total_leads or 0,
            'total_professionals': total_pros or 0,
            'total_zones': total_zones or 0,
            'leads_this_month': leads_this_month or 0,
        })
    except Exception as e:
        print(f"Error en landing_stats: {e}")
        return jsonify({
            'total_leads': 0,
            'total_professionals': 0,
            'total_zones': 0,
            'leads_this_month': 0,
        }), 500
    finally:
        if conn:
            conn.close()


@app.route('/sitemap.xml')
def sitemap():
    """Generate XML sitemap for search engines."""
    public_urls = [
        {'loc': url_for('index', _external=True), 'changefreq': 'daily', 'priority': '1.0'},
        {'loc': url_for('login', _external=True), 'changefreq': 'monthly', 'priority': '0.3'},
        {'loc': url_for('register', _external=True), 'changefreq': 'monthly', 'priority': '0.4'},
    ]
    xml = '<?xml version="1.0" encoding="UTF-8"?>\n'
    xml += '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n'
    for url in public_urls:
        xml += f'  <url>\n    <loc>{url["loc"]}</loc>\n'
        xml += f'    <changefreq>{url["changefreq"]}</changefreq>\n'
        xml += f'    <priority>{url["priority"]}</priority>\n'
        xml += '  </url>\n'
    xml += '</urlset>'
    return xml, 200, {'Content-Type': 'application/xml'}


@app.route('/robots.txt')
def robots():
    """Serve robots.txt for search engine crawlers."""
    sitemap_url = url_for('sitemap', _external=True)
    content = f"""User-agent: *
Allow: /
Disallow: /admin/
Disallow: /api/
Disallow: /login
Disallow: /register
Disallow: /usuario
Disallow: /profesional
Sitemap: {sitemap_url}
"""
    return content, 200, {'Content-Type': 'text/plain'}


@app.route('/usuario')
@login_required
def user_view():
    conn = None
    try:
        conn = get_db_connection()
        user = conn.execute('SELECT role FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        lead_count = conn.execute('SELECT COUNT(*) FROM leads WHERE user_id = ?', (session['user_id'],)).fetchone()[0]
    finally:
        if conn:
            conn.close()

    if user and user['role'] == 'professional':
        flash('Acceso denegado. Los profesionales no pueden acceder a esta sección.', 'error')
        return redirect(url_for('index'))

    return render_template('user.html', is_first_lead=lead_count == 0)


@app.route('/profesional')
@professional_required
def professional_view():
    """Muestra el panel de leads (datos se cargan dinámicamente)"""
    conn = None
    try:
        conn = get_db_connection()
        user = conn.execute('SELECT username, doc_path FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        if not user:
            return redirect(url_for('index'))

        professional = conn.execute('SELECT status FROM professionals WHERE name = ?', (user['username'],)).fetchone()
        if not professional or professional['status'] != 'approved':
            return render_template('professional.html', pending=True, doc_path=user['doc_path'])

        return render_template('professional.html', pending=False, doc_path=user['doc_path'])
    finally:
        if conn:
            conn.close()


@app.route('/profesional/lead/<int:lead_id>')
@professional_required
def lead_detail(lead_id):
    conn = None
    try:
        conn = get_db_connection()
        user = conn.execute('SELECT username FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        if not user:
            return redirect(url_for('index'))

        professional = conn.execute('SELECT status FROM professionals WHERE name = ?', (user['username'],)).fetchone()
        if not professional or professional['status'] != 'approved':
            return render_template('professional.html', leads=[], pending=True)

        lead = conn.execute('SELECT * FROM leads WHERE id = ?', (lead_id,)).fetchone()
    finally:
        if conn:
            conn.close()

    if not lead:
        return redirect(url_for('professional_view'))

    lead_dict = dict(lead)
    lead_dict['timestamp'] = convert_to_argentina_time(lead_dict['timestamp'])
    phone_raw = lead_dict.get('phone') or ''
    lead_dict['phone_e164'] = utils.normalize_phone_to_e164(phone_raw)
    lead_dict['phone_is_mobile'] = bool(lead_dict['phone_e164'] and utils.is_whatsapp_capable(lead_dict['phone_e164']))
    return render_template('lead_detail.html', lead=lead_dict)


@app.route('/admin')
@admin_required
def admin_view():
    """Muestra logs de auditoría (profesionales se cargan dinámicamente)"""
    conn = None
    try:
        conn = get_db_connection()
        audit_logs = conn.execute('SELECT * FROM audit_log ORDER BY timestamp DESC').fetchall()
    finally:
        if conn:
            conn.close()

    audit_log_converted = []
    for log in audit_logs:
        log_dict = dict(log)
        log_dict['timestamp'] = convert_to_argentina_time(log_dict['timestamp'])
        audit_log_converted.append(log_dict)

    return render_template('admin.html',
                           audit_log=audit_log_converted)


# --- CONTEXT PROCESSOR: inyectar tema en todas las templates ---
@app.context_processor
def inject_theme():
    user_id = session.get('user_id')
    theme = 'light'
    if user_id:
        try:
            prefs = models.get_user_preferences(user_id)
            theme = prefs.get('theme', 'light')
        except Exception:
            pass
    return dict(user_theme=theme)


# --- RUTAS DE API (LÓGICA DE DATOS) ---


# --- RUTA DE REGISTRO ---
@app.route('/register', methods=['GET', 'POST'])
@rate_limit.check_rate_limit(limit=5, window=60)
def register():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        phone = request.form.get('phone', '').strip()
        password = request.form.get('password', '').strip()
        raw_role = request.form.get('role', 'client')
        license_number = request.form.get('license', '').strip()

        # ✅ VALIDACIÓN DE CAMPOS OBLIGATORIOS
        if not username or len(username) < 3 or len(username) > 30:
            flash('El nombre de usuario debe tener entre 3 y 30 caracteres.', 'error')
            return redirect(url_for('register'))

        if not re.match(r'^[a-zA-Z0-9_]+$', username):
            flash('El usuario solo puede contener letras, números y guión bajo.', 'error')
            return redirect(url_for('register'))

        if not email:
            flash('El email es requerido.', 'error')
            return redirect(url_for('register'))

        is_valid_email_result, email_error = validators.validate_email(email)
        if not is_valid_email_result:
            flash(email_error, 'error')
            return redirect(url_for('register'))
        
        if not password or len(password) < 6:
            flash('La contraseña debe tener al menos 6 caracteres.', 'error')
            return redirect(url_for('register'))

        if not re.search(r'[A-Za-z]', password) or not re.search(r'[0-9]', password):
            flash('La contraseña debe contener al menos una letra y un número.', 'error')
            return redirect(url_for('register'))

        # 📞 VALIDACIÓN DE TELÉFONO
        if not phone:
            flash('El teléfono es requerido.', 'error')
            return redirect(url_for('register'))
        is_valid_phone, phone_error = validators.validate_phone(phone)
        if not is_valid_phone:
            flash(phone_error, 'error')
            return redirect(url_for('register'))

        # 🛡️ VALIDACIÓN DE SEGURIDAD CRÍTICA (Backend)
        # Detectar intentos de inyectar 'admin' o roles no autorizados
        if raw_role == 'admin':
            print(f"⚠️ ALERTA DE SEGURIDAD: Intento de registro ilegal como admin por {username}")
            flash('Acceso denegado. Solo administradores pueden asignarse ese rol.', 'error')
            return redirect(url_for('register'))
        
        # Solo permitimos roles explícitamente definidos
        if raw_role in ['client', 'professional']:
            role = raw_role
        else:
            role = 'client'
        
        # ✅ VALIDACIÓN: Profesional requiere matrícula
        if role == 'professional' and not license_number:
            flash('El número de matrícula es requerido para profesionales.', 'error')
            return redirect(url_for('register'))

        if role == 'professional' and (len(license_number) < 3 or len(license_number) > 50):
            flash('El número de matrícula debe tener entre 3 y 50 caracteres.', 'error')
            return redirect(url_for('register'))

        if role == 'professional' and not re.match(r'^[a-zA-Z0-9\-]+$', license_number):
            flash('El número de matrícula contiene caracteres no válidos.', 'error')
            return redirect(url_for('register'))

        conn = get_db_connection()
        try:
            # 1. Crear usuario con email separado, rol validado y teléfono
            cursor = conn.execute('INSERT INTO users (username, email, hash, role, phone, phone_format_valid) VALUES (?, ?, ?, ?, ?, 1)',
                                 (username, email, generate_password_hash(password), role, phone))
            
            # 2. Si es profesional, vincular con user_id real
            if role == 'professional':
                new_user_id = cursor.lastrowid
                conn.execute('INSERT INTO professionals (user_id, name, license, specialty, status) VALUES (?, ?, ?, ?, ?)',
                             (new_user_id, username, license_number, 'General', 'pending'))
            
            conn.commit()
            flash('Registro exitoso. Por favor, inicia sesión.', 'success')
            return redirect(url_for('login'))

        except sqlite3.IntegrityError as e:
            flash('El nombre de usuario ya está en uso. Por favor, elige otro.', 'error')
            return redirect(url_for('register'))
        except Exception as e:
            print(f"Error al registrar usuario: {e}")
            flash('Error al registrar. Por favor, intenta de nuevo.', 'error')
            return redirect(url_for('register'))
        finally:
            conn.close()

    return render_template('register.html')


# --- RUTA DE LOGIN ---
@app.route('/login', methods=['GET', 'POST'])
@rate_limit.check_rate_limit(limit=20, window=60)
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        conn = None
        try:
            conn = get_db_connection()
            user = conn.execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()
        finally:
            if conn:
                conn.close()

        if user and check_password_hash(user['hash'], password):
            if not user['is_active']:
                flash('Tu cuenta ha sido dada de baja. Contactá al administrador para más información.', 'error')
                return redirect(url_for('login'))

            session.clear()
            session.permanent = True
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['email'] = user['email']
            session['role'] = user['role']

            # Registrar sesión en historial
            try:
                conn2 = get_db_connection()
                conn2.execute(
                    'INSERT INTO user_login_history (user_id, ip_address, user_agent) VALUES (?, ?, ?)',
                    (user['id'], request.remote_addr or '', request.user_agent.string[:255] if request.user_agent else '')
                )
                conn2.commit()
                conn2.close()
            except Exception:
                pass

            # Remember-me: si el form lo pidió, generar token + cookie firmada
            remember_response = None
            if request.form.get('remember') == 'on':
                try:
                    selector, validator, validator_hash = utils.generate_remember_token()
                    conn3 = get_db_connection()
                    conn3.execute(
                        'INSERT INTO remember_tokens (user_id, selector, validator_hash, expires_at, ip_address, user_agent) '
                        'VALUES (?, ?, ?, ?, ?, ?)',
                        (
                            user['id'],
                            selector,
                            validator_hash,
                            utils.remember_expires_at().isoformat(),
                            request.remote_addr or '',
                            request.user_agent.string[:255] if request.user_agent else '',
                        )
                    )
                    conn3.commit()
                    conn3.close()
                    remember_response = redirect(url_for(
                        'admin_view' if user['role'] == 'admin' else
                        'professional_view' if user['role'] == 'professional' else
                        'user_view'
                    ))
                    remember_response.set_cookie(
                        config.REMEMBER_COOKIE_NAME,
                        f'{selector}:{validator}',
                        max_age=utils.remember_cookie_max_age(),
                        httponly=True,
                        secure=config.REMEMBER_COOKIE_SECURE,
                        samesite='Lax',
                        path='/',
                    )
                    utils.log_action(
                        "Remember me activado",
                        f"user_id={user['id']}, selector={selector[:8]}...",
                        session
                    )
                except Exception as e:
                    print(f"Error al crear remember token: {e}")
                    remember_response = None

            if user['role'] == 'admin':
                target = url_for('admin_view')
            elif user['role'] == 'professional':
                target = url_for('professional_view')
            else:
                target = url_for('user_view')

            if remember_response is not None:
                return remember_response
            return redirect(target)

        flash('Credenciales inválidas. Intente de nuevo.', 'error')
        return redirect(url_for('login'))

    return render_template('login.html')

# --- RUTA DE LOGOUT ---
@app.route('/logout')
def logout():
    raw_cookie = request.cookies.get(config.REMEMBER_COOKIE_NAME)
    if raw_cookie and ':' in raw_cookie:
        selector = raw_cookie.split(':', 1)[0]
        if selector:
            utils.revoke_remember_token(selector)
    session.clear()
    response = redirect(url_for('index'))
    response.delete_cookie(config.REMEMBER_COOKIE_NAME, path='/')
    return response


# --- CHECK USERNAME AVAILABILITY (rate-limited, 200 siempre) ---
@app.route('/api/auth/check-username', methods=['GET'])
@rate_limit.check_rate_limit(limit=10, window=60)
def api_check_username():
    """
    Indica si un username está disponible. Devuelve siempre 200 (no permite
    enumeración por status code). Rate-limited a 10/min por IP.
    Respuesta: {available: bool, reason: 'ok'|'taken'|'invalid'}
    """
    import random
    q = (request.args.get('q') or '').strip()
    # Latencia constante para no exponer timing
    if not (3 <= len(q) <= 30) or not re.match(r'^[a-zA-Z0-9_]+$', q):
        result = {"available": False, "reason": "invalid"}
    else:
        conn = None
        try:
            conn = get_db_connection()
            row = conn.execute('SELECT 1 FROM users WHERE username = ?', (q,)).fetchone()
            result = {"available": row is None, "reason": "ok" if row is None else "taken"}
        except Exception as e:
            print(f"Error en check-username: {e}")
            result = {"available": False, "reason": "invalid"}
        finally:
            if conn:
                conn.close()

    import time as _t
    _t.sleep(random.uniform(0.02, 0.08))
    return jsonify(result)


@app.route('/api/submit', methods=['POST'])
@rate_limit.check_rate_limit(limit=10, window=60)
def submit_lead():
    """
    Envía una solicitud de propiedad.
    Solo usuarios autenticados pueden usar este endpoint.
    """
    data = request.json
    user_id = session.get('user_id')

    if not user_id:
        return jsonify({
            "status": "error",
            "message": "Debes estar registrado para enviar solicitudes."
        }), 401

    conn = None
    try:
        conn = get_db_connection()
        user = conn.execute('SELECT id, username FROM users WHERE id = ?', (user_id,)).fetchone()

        if not user:
            return jsonify({"status": "error", "message": "Sesión no válida"}), 401

        email = session.get('email') or data.get('email', '')
        is_valid, error = validators.validate_email(email)
        if not is_valid:
            return jsonify({"status": "error", "message": error}), 400

        phone = data.get('phone', '').strip()

        # Teléfono opcional para administradores
        phone_format_valid = 0
        if session.get('role') != 'admin':
            if not phone:
                return jsonify({"status": "error", "message": "Telefono es obligatorio"}), 400
            is_valid, error = validators.validate_phone(phone)
            if not is_valid:
                return jsonify({"status": "error", "message": error}), 400
            phone_format_valid = 1

        budget = data.get('budget')
        if budget:
            is_valid, error = validators.validate_budget(budget)
            if not is_valid:
                return jsonify({"status": "error", "message": error}), 400

        zone = data.get('zone')
        if zone:
            is_valid, error = validators.validate_zone(zone)
            if not is_valid:
                return jsonify({"status": "error", "message": error}), 400

        # Validar campos requeridos
        lead_type = data.get('type')
        if not lead_type:
            return jsonify({"status": "error", "message": "El tipo de operación es requerido."}), 400
        
        zone = data.get('zone')
        if not zone:
            return jsonify({"status": "error", "message": "La zona es requerida."}), 400
        
        budget = data.get('budget')
        if not budget:
            return jsonify({"status": "error", "message": "El presupuesto es requerido."}), 400
        
        property_type = data.get('property_type', 'departamento')
        VALID_PROPERTY_TYPES = ['departamento', 'casa', 'duplex', 'penthouse', 'local_comercial']
        if property_type not in VALID_PROPERTY_TYPES:
            return jsonify({"status": "error", "message": "Tipo de propiedad no válido."}), 400

        VALID_CURRENCIES = ['ARG', 'USD', 'EUR']
        currency = data.get('currency', 'ARG')
        if currency not in VALID_CURRENCIES:
            return jsonify({"status": "error", "message": "Moneda no válida."}), 400

        VALID_LEAD_TYPES = ['Comprar Propiedad', 'Remodelación Integral', 'Construir desde Cero']
        if lead_type not in VALID_LEAD_TYPES:
            return jsonify({"status": "error", "message": "Tipo de operación no válido."}), 400

        try:
            land_area = int(data.get('land_area') or 0)
            built_area = int(data.get('built_area') or 0)
        except (ValueError, TypeError):
            land_area = 0
            built_area = 0

        if property_type == 'casa' and built_area > land_area:
            return jsonify({"status": "error", "message": "Los metros construidos no pueden ser mayores que los metros de terreno."}), 400

        province = data.get('province', '')

        conn.execute('''
            INSERT INTO leads (
                type, property_type, zone, province, budget, currency,
                phone, email, user_id, floor_block, usable_m2, elevator,
                land_area, built_area, pool, architectural_style,
                bedrooms, bathrooms, total_area, amenities,
                ambientes, parking, orientation, property_condition, property_age,
                phone_format_valid
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data.get('type'),
            property_type,
            zone,
            province,
            budget,
            currency,
            data.get('phone'),
            email,
            user_id,
            data.get('floor_block', ''),
            data.get('usable_m2', 0),
            data.get('elevator', ''),
            data.get('land_area', 0),
            data.get('built_area', 0),
            data.get('pool', ''),
            data.get('architectural_style', ''),
            data.get('bedrooms', 0),
            data.get('bathrooms', 0),
            data.get('total_area', 0),
            data.get('amenities', ''),
            data.get('ambientes', 0),
            data.get('parking', ''),
            data.get('orientation', ''),
            data.get('property_condition', ''),
            data.get('property_age', ''),
            phone_format_valid,
        ))
        conn.commit()

        return jsonify({
            "status": "success",
            "message": "Solicitud enviada con éxito. Los profesionales se contactarán contigo."
        })

    except Exception as e:
        print(f"Error en BD: {e}")
        return jsonify({"status": "error", "message": "Error al procesar la solicitud."}), 500
    finally:
        if conn:
            conn.close()



@app.route('/api/leads/filter-options')
@professional_required
def get_leads_filter_options():
    """Retorna valores distintos para poblar los filtros dinámicamente."""
    cached = filter_cache.get('filter_options')
    if cached:
        return jsonify(cached)

    conn = None
    try:
        conn = get_db_connection()
        types      = [r[0] for r in conn.execute('SELECT DISTINCT type FROM leads WHERE type IS NOT NULL ORDER BY type').fetchall()]
        prop_types = [r[0] for r in conn.execute('SELECT DISTINCT property_type FROM leads WHERE property_type IS NOT NULL ORDER BY property_type').fetchall()]
        currencies = [r[0] for r in conn.execute('SELECT DISTINCT currency FROM leads WHERE currency IS NOT NULL ORDER BY currency').fetchall()]
        zones      = [r[0] for r in conn.execute('SELECT DISTINCT zone FROM leads WHERE zone IS NOT NULL ORDER BY zone').fetchall()]

        result = {
            'types':          types,
            'property_types': prop_types,
            'currencies':     currencies,
            'zones':          zones,
        }
        filter_cache.set('filter_options', result)
        return jsonify(result)
    finally:
        if conn:
            conn.close()


@app.route('/api/leads/filter-options/invalidate', methods=['POST'])
@admin_required
def invalidate_filter_cache():
    """Invalida la caché de opciones de filtro."""
    filter_cache.invalidate()
    return jsonify({"status": "success", "message": "Caché invalidada"})


@app.route('/api/leads/stats', methods=['GET'])
def budget_stats():
    """Retorna estadísticas de presupuesto en formato JSON"""
    stats = get_budget_stats_from_db()
    return jsonify(stats)


@app.route('/api/budget-stats', methods=['GET'])
def budget_stats_for_popup():
    """Retorna estadísticas de presupuesto para el popup de frontend"""
    return jsonify({
        'min': 0,
        'max': 10000000000,
        'ranges': [],
        'currency_options': ['ARG', 'USD', 'EUR'],
    })


@app.route('/api/leads/export')
@professional_required
def export_leads_csv():
    """Genera y descarga un archivo CSV con todos los leads"""
    conn = None
    leads = []
    try:
        conn = get_db_connection()
        user = conn.execute('SELECT username FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        if not user:
            return "Acceso denegado", 403

        professional = conn.execute('SELECT status FROM professionals WHERE name = ?', (user['username'],)).fetchone()
        if not professional or professional['status'] != 'approved':
            return "Cuenta pendiente de aprobación", 403

        leads = conn.execute('SELECT id, type, zone, budget, currency, timestamp FROM leads ORDER BY timestamp DESC').fetchall()
    finally:
        if conn:
            conn.close()

    def generate():
        data = StringIO()
        writer = csv.writer(data)

        writer.writerow(['ID', 'Tipo Operacion', 'Zona', 'Presupuesto', 'Moneda', 'Fecha Registro (Argentina)'])
        yield data.getvalue()
        data.seek(0)
        data.truncate(0)

        for lead in leads:
            timestamp_argentina = convert_to_argentina_time(lead['timestamp'])
            writer.writerow([lead['id'], lead['type'], lead['zone'], lead['budget'], lead['currency'], timestamp_argentina])
            yield data.getvalue()
            data.seek(0)
            data.truncate(0)

    filename = f"leads_archestate_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return Response(
        generate(),
        mimetype='text/csv',
        headers={
            'Content-Disposition': f'attachment; filename="{filename}"'
        }
    )

@app.route('/api/leads/export/xlsx')
@professional_required
def export_leads_xlsx():
    """Genera y descarga un archivo XLSX con todos los leads"""
    conn = None
    leads = []
    try:
        conn = get_db_connection()
        user = conn.execute('SELECT username FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        if not user:
            return "Acceso denegado", 403

        professional = conn.execute('SELECT status FROM professionals WHERE name = ?', (user['username'],)).fetchone()
        if not professional or professional['status'] != 'approved':
            return "Cuenta pendiente de aprobación", 403

        leads = conn.execute('SELECT id, type, zone, budget, currency, timestamp FROM leads ORDER BY timestamp DESC').fetchall()
    finally:
        if conn:
            conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    headers = ['ID', 'Tipo Operacion', 'Zona', 'Presupuesto', 'Moneda', 'Fecha Registro (Argentina)']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    for row_num, lead in enumerate(leads, 2):
        timestamp_argentina = convert_to_argentina_time(lead['timestamp'])
        ws.cell(row=row_num, column=1, value=lead['id'])
        ws.cell(row=row_num, column=2, value=lead['type'])
        ws.cell(row=row_num, column=3, value=lead['zone'])
        ws.cell(row=row_num, column=4, value=lead['budget'])
        ws.cell(row=row_num, column=5, value=lead['currency'])
        ws.cell(row=row_num, column=6, value=timestamp_argentina)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"leads_archestate_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# /api/lead/<id>/phone ahora vive en routes.lead_bp (Fase D)


@app.route('/api/lead/<int:lead_id>/download')
@professional_required
def download_lead_pdf(lead_id):
    """Genera un PDF con los detalles del lead para descarga."""
    conn = None
    try:
        conn = get_db_connection()
        user = conn.execute('SELECT username FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        if not user:
            return "Acceso denegado", 403

        professional = conn.execute('SELECT status FROM professionals WHERE name = ?', (user['username'],)).fetchone()
        if not professional or professional['status'] != 'approved':
            return "Cuenta pendiente de aprobación", 403

        lead = conn.execute('SELECT * FROM leads WHERE id = ?', (lead_id,)).fetchone()
    finally:
        if conn:
            conn.close()

    if not lead:
        return jsonify({"status": "error", "message": "Lead no encontrado"}), 404

    lead = dict(lead)

    def pdf_safe(value):
        """Convert values to ASCII-only text safe for FPDF"""
        if value is None:
            return ''
        text = str(value)
        replacements = {
            '\u20ac': 'EUR',
            '\u00a3': 'GBP',
            '\u00a5': 'JPY',
            '\u2014': '-',
            '\u2013': '-',
            '\u2022': '-',
            '\u221a': 'sqrt',
            '\u00d7': 'x',
            '\u00f7': '/',
            '\u2122': 'TM',
            '\u00a9': '(c)',
            '\u00ae': '(R)',
            '\u2026': '...',
            '\u00b2': '2',
            '\u00b3': '3',
            '\u00b0': 'deg',
        }
        for old, new in replacements.items():
            text = text.replace(old, new)
        accents = {
            '\u00e1': 'a', '\u00e9': 'e', '\u00ed': 'i', '\u00f3': 'o', '\u00fa': 'u',
            '\u00e0': 'a', '\u00e8': 'e', '\u00ec': 'i', '\u00f2': 'o', '\u00f9': 'u',
            '\u00e4': 'a', '\u00eb': 'e', '\u00ef': 'i', '\u00f6': 'o', '\u00fc': 'u',
            '\u00e3': 'a', '\u00f5': 'o', '\u00f1': 'n',
            '\u00c1': 'A', '\u00c9': 'E', '\u00cd': 'I', '\u00d3': 'O', '\u00da': 'U',
            '\u00c0': 'A', '\u00c8': 'E', '\u00cc': 'I', '\u00d2': 'O', '\u00d9': 'U',
            '\u00c4': 'A', '\u00cb': 'E', '\u00cf': 'I', '\u00d6': 'O', '\u00dc': 'U',
            '\u00c3': 'A', '\u00d5': 'O', '\u00d1': 'N',
            '\u00e7': 'c', '\u00c7': 'C',
            '\u00df': 'ss',
        }
        for old, new in accents.items():
            text = text.replace(old, new)
        return ''.join(c if ord(c) < 128 else '?' for c in text)

    def pdf_val(value, default='-'):
        """Return safe text or default for display"""
        text = pdf_safe(value)
        return text if text else default

    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    midnight = (0, 4, 16)
    gold = (115, 90, 58)

    pdf.set_font('Times', 'BI', 20)
    pdf.set_text_color(*midnight)
    pdf.cell(0, 15, 'ArchEstate - Detalle de Lead', ln=True, align='C')
    pdf.ln(5)

    pdf.set_font('Helvetica', '', 10)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 8, f'Lead #{lead["id"]} - Informacion completa enviada por el cliente', ln=True, align='C')
    pdf.ln(10)

    def section_header(title):
        pdf.set_fill_color(*gold)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font('Helvetica', 'B', 12)
        pdf.cell(0, 8, title.upper(), ln=True, fill=True)
        pdf.set_text_color(*midnight)
        pdf.set_font('Helvetica', '', 10)
        pdf.ln(2)

    section_header('Tipo de Operacion')
    pdf.cell(0, 6, pdf_val(lead['type']), ln=True)

    section_header('Zona Geografica')
    pdf.cell(0, 6, pdf_val(lead['zone']), ln=True)

    section_header('Presupuesto')
    budget_symbol = 'USD' if lead['currency'] == 'USD' else 'EUR' if lead['currency'] == 'EUR' else '$'
    pdf.cell(0, 6, f"{budget_symbol} {pdf_val(lead['budget'])}", ln=True)

    section_header('Estilo Arquitectonico')
    pdf.cell(0, 6, pdf_val(lead.get('architectural_style'), 'No especificado'), ln=True)

    section_header('Contacto Directo')
    pdf.cell(0, 6, f"Email: {pdf_val(lead['email'])}", ln=True)
    pdf.cell(0, 6, f"Telefono: {pdf_val(lead['phone'])}", ln=True)

    section_header('Registrado')
    pdf.cell(0, 6, pdf_val(convert_to_argentina_time(lead['timestamp'])), ln=True)
    pdf.ln(5)

    section_header('Especificaciones Tecnicas')

    pdf.set_font('Helvetica', 'B', 10)
    pdf.cell(60, 8, 'Habitaciones:', border=1)
    pdf.cell(0, 8, pdf_val(lead['bedrooms']), ln=True, border=1)

    pdf.cell(60, 8, 'Banios:', border=1)
    pdf.cell(0, 8, pdf_val(lead['bathrooms']), ln=True, border=1)

    prop_type = pdf_safe(lead.get('property_type', '')).lower()
    if prop_type == 'casa':
        pdf.cell(60, 8, 'Metros de Terreno:', border=1)
        pdf.cell(0, 8, f"{pdf_val(lead['land_area'])} m2" if lead.get('land_area') else '-', ln=True, border=1)
    else:
        pdf.cell(60, 8, 'Metros Utiles:', border=1)
        pdf.cell(0, 8, f"{pdf_val(lead['usable_m2'])} m2" if lead.get('usable_m2') else '-', ln=True, border=1)

    pdf.ln(5)

    section_header('Extras y Comodidades')
    amenities = lead.get('amenities', '')
    if amenities and str(amenities).strip():
        for amenity in pdf_safe(amenities).split(','):
            stripped = amenity.strip()
            if stripped:
                pdf.cell(0, 6, f"- {stripped}", ln=True)
    else:
        pdf.cell(0, 6, 'No especificadas', ln=True)

    if prop_type == 'departamento':
        section_header('Detalles del Departamento')
        pdf.cell(0, 6, f"Piso / Bloque: {pdf_val(lead.get('floor_block'), 'No especificado')}", ln=True)
        pdf.cell(0, 6, f"Metros Utiles: {pdf_val(lead.get('usable_m2'), 'No especificado')} m2", ln=True)
        pdf.cell(0, 6, f"Ascensor: {pdf_val(lead.get('elevator'), 'No especificado')}", ln=True)
    else:
        section_header('Detalles de la Propiedad')
        pdf.cell(0, 6, f"Superficie de Terreno: {pdf_val(lead.get('land_area'), 'No especificado')} m2", ln=True)
        pdf.cell(0, 6, f"Superficie Construida: {pdf_val(lead.get('built_area'), 'No especificado')} m2", ln=True)
        pdf.cell(0, 6, f"Piscina: {pdf_val(lead.get('pool'), 'No especificado')}", ln=True)

    # Generar el PDF
    pdf_output = pdf.output(dest='S')
    if isinstance(pdf_output, str):
        pdf_output = pdf_output.encode('latin-1')

    # Crear buffer con los bytes del PDF
    buffer = io.BytesIO(pdf_output)
    buffer.seek(0)
    
    filename = f"lead_{lead['id']}.pdf"
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/pdf'
    )


@app.route('/api/lead/<int:lead_id>/toggle-status', methods=['POST'])
@professional_required
def toggle_lead_status(lead_id):
    """Toggle 'visto' o 'contactado' para un lead (por profesional actual)."""
    conn = None
    try:
        conn = get_db_connection()

        # Validar que el profesional esta aprobado (mismo patron que get_leads_api)
        user = conn.execute('SELECT username FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        if not user:
            return jsonify({'error': 'Acceso denegado'}), 403

        professional = conn.execute(
            'SELECT status FROM professionals WHERE name = ?',
            (user['username'],)
        ).fetchone()
        if not professional or professional['status'] != 'approved':
            return jsonify({'error': 'Cuenta pendiente de aprobacion'}), 403

        # Validar que el lead existe
        lead = conn.execute('SELECT id FROM leads WHERE id = ?', (lead_id,)).fetchone()
        if not lead:
            return jsonify({'error': 'Lead no encontrado'}), 404

        data = request.get_json()
        status_type = data.get('status')

        if status_type not in ('seen', 'contacted'):
            return jsonify({'error': 'Tipo de estado invalido'}), 400

        professional_id = session['user_id']
        argentina_tz = pytz.timezone('America/Argentina/Buenos_Aires')
        now = datetime.now(argentina_tz).strftime('%Y-%m-%d %H:%M:%S')

        # Buscar registro existente
        tracking = conn.execute(
            'SELECT * FROM lead_tracking WHERE professional_id = ? AND lead_id = ?',
            (professional_id, lead_id)
        ).fetchone()

        if tracking:
            # Toggle: invertir el valor actual
            current_value = tracking[status_type]
            new_value = 0 if current_value else 1
            timestamp_col = f'{status_type}_at'
            timestamp_value = now if new_value else None

            conn.execute(
                f'UPDATE lead_tracking SET {status_type} = ?, {timestamp_col} = ? WHERE professional_id = ? AND lead_id = ?',
                (new_value, timestamp_value, professional_id, lead_id)
            )
        else:
            # Crear nuevo registro con el estado activado
            seen_val = 1 if status_type == 'seen' else 0
            contacted_val = 1 if status_type == 'contacted' else 0
            seen_at = now if status_type == 'seen' else None
            contacted_at = now if status_type == 'contacted' else None

            conn.execute(
                'INSERT INTO lead_tracking (professional_id, lead_id, seen, contacted, seen_at, contacted_at) VALUES (?, ?, ?, ?, ?, ?)',
                (professional_id, lead_id, seen_val, contacted_val, seen_at, contacted_at)
            )
            new_value = 1

        conn.commit()

        return jsonify({
            'success': True,
            'status': status_type,
            'value': new_value,
            'timestamp': now if new_value else None
        })
    except Exception as e:
        print(f"Error en toggle_lead_status: {e}")
        return jsonify({'error': 'Error interno'}), 500
    finally:
        if conn:
            conn.close()


@app.route('/api/lead/<int:lead_id>/report', methods=['POST'])
@professional_required
def report_lead(lead_id):
    """Reportar un lead como telefono inexistente."""
    conn = None
    try:
        conn = get_db_connection()

        user = conn.execute('SELECT username FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        if not user:
            return jsonify({'error': 'Acceso denegado'}), 403

        professional = conn.execute(
            'SELECT status FROM professionals WHERE name = ?',
            (user['username'],)
        ).fetchone()
        if not professional or professional['status'] != 'approved':
            return jsonify({'error': 'Cuenta pendiente de aprobacion'}), 403

        lead = conn.execute('SELECT id, type, phone FROM leads WHERE id = ?', (lead_id,)).fetchone()
        if not lead:
            return jsonify({'error': 'Lead no encontrado'}), 404

        data = request.get_json() or {}
        notes = utils.safe_text(data.get('notes', ''))[:500]

        existing = conn.execute(
            'SELECT id FROM lead_reports WHERE lead_id = ? AND reported_by = ? AND status = ?',
            (lead_id, session['user_id'], 'pending')
        ).fetchone()
        if existing:
            return jsonify({'error': 'Ya reportaste este pedido anteriormente'}), 400

        conn.execute(
            'INSERT INTO lead_reports (lead_id, reported_by, reason, notes, status) VALUES (?, ?, ?, ?, ?)',
            (lead_id, session['user_id'], 'telefono_inexistente', notes, 'pending')
        )
        conn.commit()

        utils.log_action("Reporte de Lead", f"Lead ID: {lead_id} (Telefono: {lead['phone']}) reportado por {user['username']}", session)

        return jsonify({
            'success': True,
            'message': 'Pedido reportado correctamente'
        })
    except Exception as e:
        print(f"Error en report_lead: {e}")
        return jsonify({'error': 'Error interno'}), 500
    finally:
        if conn:
            conn.close()


# --- NUEVA API: OBTENER LEADS DINÁMICAMENTE ---
@app.route('/api/leads')
@professional_required
def get_leads_api():
    """API para obtener leads dinámicamente con filtros opcionales"""
    conn = None
    try:
        conn = get_db_connection()

        user = conn.execute('SELECT username FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        if not user:
            return jsonify({"status": "error", "message": "Acceso denegado"}), 403

        professional = conn.execute('SELECT status FROM professionals WHERE name = ?', (user['username'],)).fetchone()
        if not professional or professional['status'] != 'approved':
            return jsonify({"status": "error", "message": "Cuenta pendiente de aprobación"}), 403

        search         = request.args.get('search', '').strip()
        type_filter    = request.args.get('type', '').strip()
        prop_type      = request.args.get('property_type', '').strip()
        zone_filter    = request.args.get('zone', '').strip()
        min_budget     = request.args.get('min_budget', '').strip()
        max_budget     = request.args.get('max_budget', '').strip()
        budget_range   = request.args.get('budget_range', '').strip()
        currency_filter = request.args.get('currency', '').strip()
        sort_by        = request.args.get('sort', 'timestamp')
        sort_order     = request.args.get('order', 'desc')

        BUDGET_RANGES = {
            'hasta_200k':    (0,       200000),
            '200k_500k':     (200000,  500000),
            '500k_1m':       (500000,  1000000),
            '1m_2m':         (1000000, 2000000),
            'mas_2m':        (2000000, None),
        }
        if budget_range and budget_range in BUDGET_RANGES:
            rng = BUDGET_RANGES[budget_range]
            min_budget = str(rng[0]) if rng[0] else ''
            max_budget = str(rng[1]) if rng[1] else ''

        query = 'SELECT * FROM leads WHERE 1=1'
        params = []

        if search:
            query += ' AND (zone LIKE ? OR email LIKE ? OR type LIKE ? OR budget LIKE ?)'
            search_param = f'%{search}%'
            params.extend([search_param, search_param, search_param, search_param])

        if type_filter:
            query += ' AND type = ?'
            params.append(type_filter)

        if prop_type:
            query += ' AND property_type = ?'
            params.append(prop_type)

        if zone_filter:
            query += ' AND zone LIKE ?'
            params.append(f'%{zone_filter}%')

        if min_budget:
            try:
                min_val = float(min_budget)
                query += " AND CAST(REPLACE(REPLACE(budget, '.', ''), ',', '') AS REAL) >= ?"
                params.append(min_val)
            except ValueError:
                pass

        if max_budget:
            try:
                max_val = float(max_budget)
                query += " AND CAST(REPLACE(REPLACE(budget, '.', ''), ',', '') AS REAL) <= ?"
                params.append(max_val)
            except ValueError:
                pass

        if currency_filter:
            query += ' AND currency = ?'
            params.append(currency_filter)

        valid_sort_fields = ['id', 'type', 'zone', 'budget', 'timestamp', 'email']
        if sort_by not in valid_sort_fields:
            sort_by = 'timestamp'

        order = 'DESC' if sort_order.lower() == 'desc' else 'ASC'
        if order not in ('ASC', 'DESC'):
            order = 'ASC'
        query += f' ORDER BY {sort_by} {order}, id DESC'

        leads = conn.execute(query, params).fetchall()

        leads_list = []
        for lead in leads:
            lead_dict = dict(lead)
            lead_dict['timestamp'] = convert_to_argentina_time(lead_dict['timestamp'])
            phone_raw = lead_dict.get('phone') or ''
            phone_e164 = utils.normalize_phone_to_e164(phone_raw)
            lead_dict['phone_e164'] = phone_e164
            lead_dict['phone_is_mobile'] = bool(phone_e164 and utils.is_whatsapp_capable(phone_e164))
            leads_list.append(lead_dict)

        # Obtener tracking del profesional actual para estos leads
        professional_id = session['user_id']
        lead_ids = [lead['id'] for lead in leads_list]

        tracking_map = {}
        if lead_ids:
            placeholders = ','.join(['?'] * len(lead_ids))
            tracking_rows = conn.execute(
                f'SELECT lead_id, seen, contacted FROM lead_tracking WHERE professional_id = ? AND lead_id IN ({placeholders})',
                [professional_id] + lead_ids
            ).fetchall()
            for row in tracking_rows:
                tracking_map[row['lead_id']] = {
                    'seen': bool(row['seen']),
                    'contacted': bool(row['contacted'])
                }

        # Agregar tracking a cada lead
        for lead in leads_list:
            tracking = tracking_map.get(lead['id'], {'seen': False, 'contacted': False})
            lead['tracking'] = tracking

        return jsonify({
            "success": True,
            "leads": leads_list,
            "total": len(leads_list)
        })
    except Exception as e:
        print(f"Error en get_leads_api: {e}")
        return jsonify({"status": "error", "message": "Error interno del servidor"}), 500
    finally:
        if conn:
            conn.close()


# --- NUEVA API: OBTENER PROFESIONALES DINÁMICAMENTE ---
@app.route('/api/professionals')
@admin_required
def get_professionals_api():
    """API para obtener profesionales dinámicamente con filtros"""
    conn = None
    try:
        conn = get_db_connection()

        search = request.args.get('search', '').strip()
        status_filter = request.args.get('status', '').strip()
        specialty_filter = request.args.get('specialty', '').strip()
        sort_by = request.args.get('sort', 'id')
        sort_order = request.args.get('order', 'desc')

        query = '''
            SELECT p.*,
                   u.doc_path,
                   u.id   AS user_id,
                   u.is_active
            FROM professionals p
            LEFT JOIN users u ON (
                (p.user_id IS NOT NULL AND p.user_id = u.id)
                OR
                (p.user_id IS NULL AND p.name = u.username)
            )
            WHERE 1=1
        '''
        params = []

        if search:
            query += ' AND (p.name LIKE ? OR p.license LIKE ? OR p.specialty LIKE ?)'
            search_param = f'%{search}%'
            params.extend([search_param, search_param, search_param])

        if status_filter:
            query += ' AND p.status = ?'
            params.append(status_filter)

        if specialty_filter:
            query += ' AND p.specialty LIKE ?'
            params.append(f'%{specialty_filter}%')

        valid_sort_fields = ['id', 'name', 'license', 'specialty', 'status']
        if sort_by not in valid_sort_fields:
            sort_by = 'id'

        order = 'DESC' if sort_order.lower() == 'desc' else 'ASC'
        if order not in ('ASC', 'DESC'):
            order = 'ASC'
        query += f' ORDER BY p.{sort_by} {order}'

        professionals = conn.execute(query, params).fetchall()

        pros_list = []
        for pro in professionals:
            pro_dict = dict(pro)
            pros_list.append(pro_dict)

        return jsonify({
            "success": True,
            "professionals": pros_list,
            "total": len(pros_list)
        })
    except Exception as e:
        print(f"Error en get_professionals_api: {e}")
        return jsonify({"status": "error", "message": "Error interno del servidor"}), 500
    finally:
        if conn:
            conn.close()


@app.route('/api/admin/professional/<int:pro_id>/status', methods=['POST'])
@rate_limit.check_rate_limit(limit=10, window=60)
@admin_required
def update_pro_status(pro_id):
    """Actualiza el estado de un profesional en la BD y registra la acción"""
    data = request.json
    new_status = data.get('status')

    if new_status not in ['approved', 'rejected']:
        return jsonify({"status": "error", "message": "Estado no válido"}), 400

    conn = None
    try:
        conn = get_db_connection()
        pro = conn.execute('SELECT name FROM professionals WHERE id = ?', (pro_id,)).fetchone()

        if pro:
            conn.execute('UPDATE professionals SET status = ? WHERE id = ?', (new_status, pro_id))
            conn.commit()

            action = "Aprobación" if new_status == 'approved' else "Rechazo"
            utils.log_action(action, pro['name'], session)
            return jsonify({"status": "success", "message": f"Profesional {action.lower()} correctamente"})

        return jsonify({"error": "Profesional no encontrado"}), 404
    finally:
        if conn:
            conn.close()


@app.route('/api/admin/stats')
@login_required
def admin_stats():
    """Retorna estadísticas agregadas para el dashboard del admin"""
    conn = None
    try:
        conn = get_db_connection()

        # Total de leads
        total_leads = conn.execute('SELECT COUNT(*) FROM leads').fetchone()[0]

        # Leads por tipo de operación
        leads_by_type = conn.execute(
            'SELECT type, COUNT(*) as count FROM leads GROUP BY type ORDER BY count DESC'
        ).fetchall()

        # Leads por zona (top 5)
        leads_by_zone = conn.execute(
            'SELECT zone, COUNT(*) as count FROM leads GROUP BY zone ORDER BY count DESC LIMIT 5'
        ).fetchall()

        # Leads por presupuesto
        leads_by_budget = conn.execute(
            'SELECT budget, COUNT(*) as count FROM leads GROUP BY budget ORDER BY count DESC'
        ).fetchall()

        # Leads por mes (últimos 6 meses)
        leads_by_month = conn.execute('''
            SELECT strftime('%Y-%m', timestamp) as month, COUNT(*) as count
            FROM leads
            GROUP BY month
            ORDER BY month DESC
            LIMIT 6
        ''').fetchall()

        # Estado de profesionales
        pros_stats = conn.execute(
            'SELECT status, COUNT(*) as count FROM professionals GROUP BY status'
        ).fetchall()

        # Total de usuarios por rol
        users_by_role = conn.execute(
            'SELECT role, COUNT(*) as count FROM users GROUP BY role'
        ).fetchall()

        # Acciones del log de auditoría
        audit_actions = conn.execute(
            'SELECT action, COUNT(*) as count FROM audit_log GROUP BY action ORDER BY count DESC'
        ).fetchall()

        # Conteo de reportes pendientes
        pending_reports = conn.execute(
            "SELECT COUNT(*) FROM lead_reports WHERE status = 'pending'"
        ).fetchone()[0]

        return jsonify({
            'total_leads': total_leads,
            'leads_by_type': [{'label': r['type'], 'value': r['count']} for r in leads_by_type],
            'leads_by_zone': [{'label': r['zone'], 'value': r['count']} for r in leads_by_zone],
            'leads_by_budget': [{'label': r['budget'], 'value': r['count']} for r in leads_by_budget],
            'leads_by_month': [{'label': r['month'], 'value': r['count']} for r in reversed(leads_by_month)],
            'pros_stats': [{'label': r['status'], 'value': r['count']} for r in pros_stats],
            'users_by_role': [{'label': r['role'], 'value': r['count']} for r in users_by_role],
            'audit_actions': [{'label': r['action'], 'value': r['count']} for r in audit_actions],
            'pending_reports': pending_reports,
        })
    except Exception as e:
        print(f"Error en admin_stats: {e}")
        return jsonify({"error": "Error interno"}), 500
    finally:
        if conn:
            conn.close()


@app.route('/api/admin/lead/<int:lead_id>', methods=['GET'])
@admin_required
def admin_lead_detail(lead_id):
    """Obtener detalles de un lead para el admin."""
    conn = None
    try:
        conn = get_db_connection()
        lead = conn.execute('SELECT * FROM leads WHERE id = ?', (lead_id,)).fetchone()
        if not lead:
            return jsonify({'error': 'Lead no encontrado'}), 404

        lead_dict = dict(lead)
        lead_dict['timestamp'] = convert_to_argentina_time(lead_dict['timestamp'])

        return jsonify({
            'success': True,
            'lead': lead_dict
        })
    except Exception as e:
        print(f"Error en admin_lead_detail: {e}")
        return jsonify({'error': 'Error interno'}), 500
    finally:
        if conn:
            conn.close()


@app.route('/api/admin/reports', methods=['GET'])
@admin_required
def get_lead_reports():
    """Obtener todos los reportes de leads para el admin."""
    conn = None
    try:
        conn = get_db_connection()

        reports = conn.execute('''
            SELECT
                lr.id,
                lr.lead_id,
                lr.reason,
                lr.notes,
                lr.status,
                lr.reviewed_by,
                lr.reviewed_at,
                lr.created_at,
                u.username as reported_by_name,
                l.type as lead_type,
                l.zone as lead_zone,
                l.phone as lead_phone,
                l.budget as lead_budget,
                l.timestamp as lead_timestamp
            FROM lead_reports lr
            JOIN users u ON lr.reported_by = u.id
            LEFT JOIN leads l ON lr.lead_id = l.id
            ORDER BY lr.created_at DESC
        ''').fetchall()

        reports_list = []
        for r in reports:
            rd = dict(r)
            if rd['lead_timestamp']:
                rd['lead_timestamp'] = convert_to_argentina_time(rd['lead_timestamp'])
            rd['created_at'] = convert_to_argentina_time(rd['created_at'])
            reports_list.append(rd)

        status_counts = {}
        for r in conn.execute('SELECT status, COUNT(*) as c FROM lead_reports GROUP BY status').fetchall():
            status_counts[r['status']] = r['c']

        return jsonify({
            'success': True,
            'reports': reports_list,
            'total': len(reports_list),
            'status_counts': status_counts
        })
    except Exception as e:
        print(f"Error en get_lead_reports: {e}")
        return jsonify({'error': 'Error interno'}), 500
    finally:
        if conn:
            conn.close()


@app.route('/api/admin/telemetry', methods=['GET'])
@admin_required
def get_telemetry():
    """Resumen de eventos de telemetría WhatsApp/SMS para el dashboard admin (Fase F)."""
    conn = None
    try:
        conn = get_db_connection()

        period = request.args.get('period', '30d')
        days = {'7d': 7, '30d': 30, '90d': 90, '1y': 365}.get(period, 30)
        since_clause = f"datetime('now', '-{days} days')"

        event_counts = {}
        for row in conn.execute(
            f"SELECT event, COUNT(*) as c FROM events "
            f"WHERE ts >= {since_clause} GROUP BY event ORDER BY c DESC"
        ).fetchall():
            event_counts[row['event']] = row['c']

        wa_clicks = event_counts.get('wa_button_clicked', 0)
        wa_opens = event_counts.get('wa_link_generated', 0)
        wa_invalid = event_counts.get('wa_invalid_number', 0)
        sms_fallbacks = event_counts.get('sms_fallback_used', 0)
        otp_sent = event_counts.get('otp_sent', 0)
        otp_verified = event_counts.get('otp_verified', 0)
        otp_failed = event_counts.get('otp_verify_failed', 0)
        ctr = round(100 * wa_opens / wa_clicks, 1) if wa_clicks else 0.0

        consent_by_channel = {}
        for row in conn.execute(
            f"SELECT channel, COUNT(*) as c FROM consent_log "
            f"WHERE created_at >= {since_clause} GROUP BY channel"
        ).fetchall():
            consent_by_channel[row['channel']] = row['c']

        top_pros = []
        for row in conn.execute(
            f"SELECT u.username, COUNT(*) as clicks FROM events e "
            f"JOIN users u ON e.user_id = u.id "
            f"WHERE e.event = 'wa_link_generated' AND e.ts >= {since_clause} "
            f"GROUP BY u.username ORDER BY clicks DESC LIMIT 5"
        ).fetchall():
            top_pros.append({'username': row['username'], 'clicks': row['clicks']})

        return jsonify({
            'success': True,
            'period': period,
            'event_counts': event_counts,
            'metrics': {
                'wa_button_clicks': wa_clicks,
                'wa_links_generated': wa_opens,
                'wa_invalid_numbers': wa_invalid,
                'sms_fallbacks': sms_fallbacks,
                'wa_click_through_rate_pct': ctr,
                'otp_sent': otp_sent,
                'otp_verified': otp_verified,
                'otp_failed': otp_failed,
                'otp_success_rate_pct': round(100 * otp_verified / otp_sent, 1) if otp_sent else 0.0,
            },
            'consent_by_channel': consent_by_channel,
            'top_professionals': top_pros,
        })
    except Exception as e:
        print(f"Error en get_telemetry: {e}")
        return jsonify({'error': 'Error interno'}), 500
    finally:
        if conn:
            conn.close()


@app.route('/api/admin/report/<int:report_id>/delete', methods=['POST'])
@admin_required
def delete_reported_lead(report_id):
    """Eliminar (soft delete) un lead reportado como falso."""
    conn = None
    try:
        conn = get_db_connection()

        report = conn.execute(
            'SELECT * FROM lead_reports WHERE id = ?', (report_id,)
        ).fetchone()
        if not report:
            return jsonify({'error': 'Reporte no encontrado'}), 404

        if report['status'] == 'deleted':
            return jsonify({'error': 'El reporte ya esta eliminado'}), 400

        lead_id = report['lead_id']
        argentina_tz = pytz.timezone('America/Argentina/Buenos_Aires')
        now = datetime.now(argentina_tz).strftime('%Y-%m-%d %H:%M:%S')

        conn.execute(
            'UPDATE lead_reports SET status = ?, reviewed_by = ?, reviewed_at = ? WHERE id = ?',
            ('deleted', session.get('username'), now, report_id)
        )
        conn.commit()
        utils.log_action("Eliminacion de Lead", f"Lead ID: {lead_id} eliminado tras reporte #{report_id} por {session.get('username')}", session)

        return jsonify({
            'success': True,
            'message': 'Lead eliminado correctamente'
        })
    except Exception as e:
        print(f"Error en delete_reported_lead: {e}")
        return jsonify({'error': 'Error interno'}), 500
    finally:
        if conn:
            conn.close()


@app.route('/api/admin/report/<int:report_id>/dismiss', methods=['POST'])
@admin_required
def dismiss_report(report_id):
    """Descartar un reporte (marcar como revisado sin eliminar lead)."""
    conn = None
    try:
        conn = get_db_connection()

        report = conn.execute(
            'SELECT * FROM lead_reports WHERE id = ?', (report_id,)
        ).fetchone()
        if not report:
            return jsonify({'error': 'Reporte no encontrado'}), 404

        argentina_tz = pytz.timezone('America/Argentina/Buenos_Aires')
        now = datetime.now(argentina_tz).strftime('%Y-%m-%d %H:%M:%S')

        conn.execute(
            'UPDATE lead_reports SET status = ?, reviewed_by = ?, reviewed_at = ? WHERE id = ?',
            ('dismissed', session.get('username'), now, report_id)
        )
        conn.commit()
        utils.log_action("Reporte Descartado", f"Reporte #{report_id} descartado por {session.get('username')}", session)

        return jsonify({
            'success': True,
            'message': 'Reporte descartado'
        })
    except Exception as e:
        print(f"Error en dismiss_report: {e}")
        return jsonify({'error': 'Error interno'}), 500
    finally:
        if conn:
            conn.close()


@app.route('/api/admin/report/<int:report_id>/restore', methods=['POST'])
@admin_required
def restore_report(report_id):
    """Restaurar un reporte eliminado o descartado a estado pendiente."""
    conn = None
    try:
        conn = get_db_connection()

        report = conn.execute(
            'SELECT * FROM lead_reports WHERE id = ?', (report_id,)
        ).fetchone()
        if not report:
            return jsonify({'error': 'Reporte no encontrado'}), 404

        if report['status'] == 'pending':
            return jsonify({'error': 'El reporte ya esta pendiente'}), 400

        conn.execute(
            'UPDATE lead_reports SET status = ?, reviewed_by = NULL, reviewed_at = NULL WHERE id = ?',
            ('pending', report_id)
        )
        conn.commit()
        utils.log_action("Reporte Restaurado", f"Reporte #{report_id} restaurado a pendiente por {session.get('username')}", session)

        return jsonify({
            'success': True,
            'message': 'Reporte restaurado correctamente'
        })
    except Exception as e:
        print(f"Error en restore_report: {e}")
        return jsonify({'error': 'Error interno'}), 500
    finally:
        if conn:
            conn.close()


# --- ESTADO DEL DOCUMENTO DEL PROFESIONAL ---
@app.route('/api/professional/doc-status', methods=['GET'])
@login_required
def get_doc_status():
    """Retorna el estado del documento del profesional autenticado."""
    conn = None
    try:
        conn = get_db_connection()
        user = conn.execute('SELECT doc_path FROM users WHERE id = ?', (session['user_id'],)).fetchone()

        if not user:
            return jsonify({"error": "Usuario no encontrado"}), 404

        doc_path = user['doc_path']
        has_doc  = bool(doc_path)

        # Verificar que el archivo físico exista
        if has_doc:
            full_path = os.path.join(app.config['UPLOAD_FOLDER'], doc_path)
            has_doc   = os.path.exists(full_path)

        return jsonify({
            "has_doc":   has_doc,
            "filename":  doc_path if has_doc else None,
            # Nombre legible: eliminar el prefijo "user_ID_"
            "display_name": re.sub(r'^user_\d+_', '', doc_path) if has_doc and doc_path else None,
        })
    except Exception as e:
        print(f"Error en get_doc_status: {e}")
        return jsonify({"error": "Error interno"}), 500
    finally:
        if conn:
            conn.close()


# --- RUTA PARA QUE EL PROFESIONAL SUBA SU DOC ---
MAX_UPLOAD_BYTES = 10 * 1024 * 1024  # 10 MB

@app.route('/api/professional/upload', methods=['POST'])
@rate_limit.check_rate_limit(limit=5, window=60)
@professional_required
def upload_professional_doc():
    if 'document' not in request.files:
        return jsonify({"error": "No se incluyó ningún archivo en la solicitud."}), 400

    file = request.files['document']
    if not file or file.filename == '':
        return jsonify({"error": "No se seleccionó ningún archivo."}), 400

    if not allowed_file(file.filename):
        return jsonify({
            "error": "Tipo de archivo no permitido. Usá PDF, JPG o PNG."
        }), 415

    # Validar MIME type real (magic bytes)
    mime_valid, detected_ext, mime_error = utils.validate_mime_type(file, file.filename)
    if not mime_valid:
        return jsonify({"error": mime_error}), 415

    # Validar tamaño (leer en memoria para chequear)
    file.seek(0, 2)          # ir al final
    size = file.tell()
    file.seek(0)             # volver al inicio
    if size > config.MAX_UPLOAD_SIZE:
        return jsonify({"error": "El archivo supera el límite de 10 MB."}), 413

    # Nombre seguro con prefijo de usuario
    original_name = secure_filename(file.filename)
    filename      = f"user_{session['user_id']}_{original_name}"
    upload_dir    = app.config['UPLOAD_FOLDER']

    os.makedirs(upload_dir, exist_ok=True)

    # Eliminar documento anterior si existe
    conn = None
    try:
        conn = get_db_connection()
        prev_user = conn.execute('SELECT doc_path FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        if prev_user and prev_user['doc_path']:
            prev_path = os.path.join(upload_dir, prev_user['doc_path'])
            if os.path.exists(prev_path):
                try:
                    os.remove(prev_path)
                except Exception:
                    pass  # No bloquear el flujo si falla el borrado

        file.save(os.path.join(upload_dir, filename))

        conn.execute('UPDATE users SET doc_path = ? WHERE id = ?', (filename, session['user_id']))
        conn.commit()
    finally:
        if conn:
            conn.close()

    utils.log_action("Subida de Documento", f"Usuario ID: {session['user_id']}", session)

    return jsonify({
        "status":       "success",
        "message":      "Documento subido correctamente.",
        "filename":     filename,
        "display_name": original_name,
    })

# --- RUTA PARA QUE EL ADMIN DESCARGUE EL DOC ---
@app.route('/admin/download_doc/<int:user_id>')
@admin_required 
def download_professional_doc(user_id):
    conn = None
    try:
        conn = get_db_connection()
        user = conn.execute('SELECT doc_path FROM users WHERE id = ?', (user_id,)).fetchone()

        if not user or not user['doc_path']:
            return "El profesional no ha subido ningún documento aún.", 404

        # Usamos la ruta configurada arriba
        directory = app.config['UPLOAD_FOLDER']
        filename = user['doc_path']

        # Verificamos si el archivo físico realmente existe en el disco
        if not os.path.exists(os.path.join(directory, filename)):
            return f"Error: El archivo {filename} no existe en el servidor.", 404

        # as_attachment=True fuerza la descarga en lugar de abrirlo en el navegador
        return send_from_directory(directory, filename, as_attachment=True)

    except Exception as e:
        return f"Error interno: {str(e)}", 500
    finally:
        if conn:
            conn.close()


# --- RUTA PARA QUE EL PROFESIONAL DESCARGUE SU PROPIO DOC ---
@app.route('/profesional/download_doc')
@professional_required
def download_own_doc():
    conn = None
    try:
        conn = get_db_connection()
        user = conn.execute('SELECT doc_path FROM users WHERE id = ?', (session['user_id'],)).fetchone()

        if not user or not user['doc_path']:
            flash('No has subido ningún documento aún.', 'error')
            return redirect(url_for('professional_view'))

        # Usamos la ruta configurada arriba
        directory = app.config['UPLOAD_FOLDER']
        filename = user['doc_path']

        # Verificamos si el archivo físico realmente existe en el disco
        if not os.path.exists(os.path.join(directory, filename)):
            flash(f'Error: El archivo {filename} no existe en el servidor.', 'error')
            return redirect(url_for('professional_view'))

        # as_attachment=True fuerza la descarga en lugar de abrirlo en el navegador
        return send_from_directory(directory, filename, as_attachment=True)

    except Exception as e:
        flash(f'Error interno: {str(e)}', 'error')
        return redirect(url_for('professional_view'))
    finally:
        if conn:
            conn.close()

# --- GESTIÓN DE USUARIOS (ADMIN) ---

@app.route('/admin/usuarios')
@admin_required
def user_management_view():
    """Vista de gestión de usuarios para el administrador."""
    return render_template('user_management.html')


@app.route('/api/admin/users', methods=['GET'])
@admin_required
def get_all_users():
    """Retorna todos los usuarios registrados (sin exponer el hash)."""
    search      = request.args.get('search', '').strip()
    role_filter = request.args.get('role', '').strip()
    active_filter = request.args.get('active', '').strip()  # 'all' | '1' | '0'

    conn = None
    try:
        conn = get_db_connection()
        query = 'SELECT id, username, email, phone, role, is_active FROM users WHERE 1=1'
        params = []

        if search:
            query += ' AND (username LIKE ? OR email LIKE ?)'
            params += [f'%{search}%', f'%{search}%']

        if role_filter:
            query += ' AND role = ?'
            params.append(role_filter)

        if active_filter in ('0', '1'):
            query += ' AND is_active = ?'
            params.append(int(active_filter))

        query += ' ORDER BY is_active DESC, id ASC'   # activos primero
        users = conn.execute(query, params).fetchall()

        return jsonify({
            'success': True,
            'users': [dict(u) for u in users],
            'total': len(users)
        })
    except Exception as e:
        print(f"Error en get_all_users: {e}")
        return jsonify({"error": "Error interno"}), 500
    finally:
        if conn:
            conn.close()


@app.route('/api/admin/user/<int:user_id>/reset-password', methods=['POST'])
@rate_limit.check_rate_limit(limit=5, window=60)
@admin_required
def admin_reset_password(user_id):
    """Resetea la contraseña de un usuario. Solo accesible por administradores."""
    data = request.json
    new_password = (data.get('password') or '').strip()

    if not new_password or len(new_password) < 6:
        return jsonify({"error": "La contraseña debe tener al menos 6 caracteres."}), 400

    conn = None
    try:
        conn = get_db_connection()
        user = conn.execute('SELECT username, role FROM users WHERE id = ?', (user_id,)).fetchone()

        if not user:
            return jsonify({"error": "Usuario no encontrado."}), 404

        # Seguridad: no permitir resetear la contraseña de otro admin
        if user['role'] == 'admin' and user_id != session.get('user_id'):
            return jsonify({"error": "No se puede resetear la contraseña de otro administrador."}), 403

        conn.execute('UPDATE users SET hash = ? WHERE id = ?',
                     (generate_password_hash(new_password), user_id))
        conn.commit()
    finally:
        if conn:
            conn.close()

    utils.log_action("Reset de Contraseña", f"Usuario: {user['username']} (ID: {user_id})", session)

    return jsonify({
        "status": "success",
        "message": f"Contraseña de '{user['username']}' actualizada correctamente."
    })
    
    
@app.route('/api/admin/user/<int:user_id>/set-active', methods=['POST'])
@rate_limit.check_rate_limit(limit=10, window=60)
@admin_required
def admin_set_user_active(user_id):
    """Da de baja o reactiva una cuenta de usuario. Solo admins. No aplica a otros admins."""
    data      = request.json
    new_state = data.get('is_active')  # True → reactivar, False → dar de baja

    if new_state not in (True, False):
        return jsonify({"error": "Estado inválido."}), 400

    conn = None
    try:
        conn = get_db_connection()
        user = conn.execute('SELECT username, role FROM users WHERE id = ?', (user_id,)).fetchone()

        if not user:
            return jsonify({"error": "Usuario no encontrado."}), 404

        # Protección: no se puede dar de baja a otro administrador
        if user['role'] == 'admin':
            return jsonify({"error": "No se puede dar de baja a un administrador."}), 403

        # Protección: un admin no se da de baja a sí mismo
        if user_id == session.get('user_id'):
            return jsonify({"error": "No podés darte de baja a vos mismo."}), 403

        conn.execute('UPDATE users SET is_active = ? WHERE id = ?', (1 if new_state else 0, user_id))
        conn.commit()
    finally:
        if conn:
            conn.close()

    action  = "Reactivación de Cuenta" if new_state else "Baja de Cuenta"
    message = f"Usuario '{user['username']}' {'reactivado' if new_state else 'dado de baja'} correctamente."
    utils.log_action(action, f"Usuario: {user['username']} (ID: {user_id})", session)

    return jsonify({"status": "success", "message": message, "is_active": new_state})


@app.route('/api/user/update-phone', methods=['POST'])
@rate_limit.check_rate_limit(limit=10, window=60)
def update_user_phone():
    """Actualiza el teléfono del usuario logueado. Invalida OTP pendiente si el teléfono cambia."""
    if 'user_id' not in session:
        return jsonify({"error": "No autorizado"}), 401

    data = request.json
    phone = (data.get('phone') or '').strip()

    if not phone:
        return jsonify({"error": "El teléfono no puede estar vacío."}), 400

    is_valid, error = validators.validate_phone(phone)
    if not is_valid:
        return jsonify({"error": error}), 400

    conn = None
    try:
        conn = get_db_connection()
        current = conn.execute('SELECT phone, phone_verified FROM users WHERE id = ?',
                               (session['user_id'],)).fetchone()
        old_phone = current['phone'] if current else ''

        e164 = utils.normalize_phone_to_e164(phone)
        ntype = ''
        if e164:
            parsed = utils._parse_phone(e164)
            if parsed is not None:
                t = phonenumbers.number_type(parsed)
                ntype = ('mobile' if t == PhoneNumberType.MOBILE
                         else 'fixed_or_mobile' if t == PhoneNumberType.FIXED_LINE_OR_MOBILE
                         else 'fixed' if t == PhoneNumberType.FIXED_LINE
                         else 'other')

        invalidate_otp = (old_phone != phone)

        if invalidate_otp:
            conn.execute(
                'UPDATE users SET phone = ?, phone_e164 = ?, phone_number_type = ?, '
                'phone_format_valid = 1, phone_verified = 0, verification_code = \'\', verification_expires = NULL '
                'WHERE id = ?',
                (phone, e164, ntype, session['user_id'])
            )
            utils.log_event(user_id=session['user_id'], event='phone_changed',
                            props={'old_hash': utils.hash_phone_digits(old_phone),
                                   'new_hash': utils.hash_phone_digits(phone),
                                   'e164': bool(e164)}, conn=conn)
        else:
            conn.execute(
                'UPDATE users SET phone_e164 = ?, phone_number_type = ? WHERE id = ?',
                (e164, ntype, session['user_id'])
            )
        conn.commit()

        return jsonify({
            "status": "success",
            "message": "Teléfono actualizado correctamente." if not invalidate_otp
                       else "Teléfono actualizado. Vuelve a verificarlo.",
            "phone": phone,
            "phone_e164": e164,
            "phone_verified": 0 if invalidate_otp else (current['phone_verified'] if current else 0),
        })
    except Exception as e:
        print(f"Error en update_user_phone: {e}")
        return jsonify({"error": "Error al actualizar el teléfono."}), 500
    finally:
        if conn:
            conn.close()


@app.route('/api/phone/send-code', methods=['POST'])
@rate_limit.check_rate_limit(limit=3, window=60)
def send_verification_code():
    """
    Envía un código de verificación de 6 dígitos por el canal preferido del usuario
    (sms | whatsapp | auto). Refactorizado en Fase C para usar VerifierRouter.
    """
    if 'user_id' not in session:
        return jsonify({"error": "No autorizado"}), 401

    user_id = session['user_id']
    conn = get_db_connection()
    try:
        user = conn.execute(
            'SELECT username, phone, phone_e164, phone_verified FROM users WHERE id = ?',
            (user_id,)
        ).fetchone()
        if not user:
            return jsonify({"error": "Usuario no encontrado"}), 404

        phone = user['phone'] or ''
        if not phone:
            return jsonify({"error": "No tenés teléfono registrado. Guardalo en tu perfil primero."}), 400

        if user['phone_verified'] == 1:
            return jsonify({"error": "El teléfono ya está verificado."}), 400

        is_valid_phone, phone_error = validators.validate_phone(phone)
        if not is_valid_phone:
            return jsonify({"error": phone_error}), 400

        phone_e164 = user['phone_e164'] or utils.normalize_phone_to_e164(phone)
        if not phone_e164:
            return jsonify({"error": "No se pudo normalizar el teléfono a E.164."}), 400

        prefs = models.get_user_preferences(user_id)
        preferred_channel = (prefs.get('preferred_channel') or 'auto').lower()

        code = f"{secrets.randbelow(1000000):06d}"
        expires = datetime.now() + timedelta(minutes=10)

        conn.execute(
            'UPDATE users SET verification_code = ?, verification_expires = ? WHERE id = ?',
            (code, expires.isoformat(), user_id)
        )
        conn.commit()

        from services.verifier import get_default_router
        router = get_default_router()
        result = router.send_otp(phone_e164, code, preferred_channel=preferred_channel, ttl_minutes=10)

        conn.execute(
            'INSERT INTO consent_log (user_id, channel, ip, user_agent) VALUES (?, ?, ?, ?)',
            (user_id, result.channel, rate_limit.get_client_ip(), request.headers.get('User-Agent', '')[:255])
        )
        conn.commit()

        utils.log_action(
            f"Envío código verificación teléfono ({result.channel})",
            f"user={user['username']}, channel={result.channel}, phone_hash={utils.hash_phone_digits(phone_e164)}",
            session,
            conn=conn
        )
        utils.log_event(user_id=user_id, event='otp_sent',
                        props={'channel': result.channel, 'preferred': preferred_channel,
                               'phone_hash': utils.hash_phone_digits(phone_e164)},
                        conn=conn)

        return jsonify({
            "status": "success" if result.ok else "error",
            "message": result.message,
            "channel": result.channel,
            "phone_e164": phone_e164,
            "deep_link": (result.meta or {}).get('deep_link') if result.ok else None,
        })
    except Exception as e:
        print(f"Error en send_verification_code: {e}")
        return jsonify({"error": "Error al enviar el código."}), 500
    finally:
        if conn:
            conn.close()


@app.route('/api/phone/verify', methods=['POST'])
@rate_limit.check_rate_limit(limit=5, window=60)
def verify_phone_code():
    """Verifica el código OTP ingresado por el usuario."""
    if 'user_id' not in session:
        return jsonify({"error": "No autorizado"}), 401

    data = request.json
    code = (data.get('code') or '').strip()

    if not code or not code.isdigit() or len(code) != 6:
        return jsonify({"error": "Código inválido. Debe ser de 6 dígitos."}), 400

    user_id = session['user_id']
    conn = get_db_connection()
    try:
        user = conn.execute(
            'SELECT username, phone, verification_code, verification_expires, phone_verified FROM users WHERE id = ?',
            (user_id,)
        ).fetchone()
        if not user:
            return jsonify({"error": "Usuario no encontrado"}), 404

        if user['phone_verified'] == 1:
            return jsonify({"error": "El teléfono ya está verificado."}), 400

        stored_code = user['verification_code'] or ''
        expires_str = user['verification_expires'] or ''

        if not stored_code or not expires_str:
            return jsonify({"error": "No hay código pendiente. Solicitá uno nuevo."}), 400

        import datetime as dt_module
        try:
            expires = dt_module.datetime.fromisoformat(expires_str)
            if dt_module.datetime.now() > expires:
                utils.log_event(user_id=user_id, event='otp_expired', conn=conn)
                return jsonify({"error": "Código expirado. Solicitá uno nuevo."}), 410
        except ValueError:
            return jsonify({"error": "Error de validación. Solicitá un nuevo código."}), 400

        if code != stored_code:
            utils.log_action(
                "Intento fallido verificación teléfono",
                f"user={user['username']}, code_ingresado={code}",
                session,
                conn=conn
            )
            utils.log_event(user_id=user_id, event='otp_verify_failed', conn=conn)
            return jsonify({"error": "Código incorrecto."}), 400

        conn.execute(
            'UPDATE users SET phone_verified = 1, verification_code = \'\', verification_expires = NULL WHERE id = ?',
            (user_id,)
        )
        conn.commit()

        utils.log_action(
            "Telefono verificado correctamente",
            f"user={user['username']}, phone_hash={utils.hash_phone_digits(user['phone'] or '')}",
            session,
            conn=conn
        )
        utils.log_event(user_id=user_id, event='otp_verified', conn=conn)

        return jsonify({"status": "success", "message": "Teléfono verificado correctamente."})

    except Exception as e:
        print(f"Error en verify_phone_code: {e}")
        return jsonify({"error": "Error al verificar el código."}), 500
    finally:
        if conn:
            conn.close()


from routes_profile import profile_bp
app.register_blueprint(profile_bp)

from routes.lead_bp import lead_bp
app.register_blueprint(lead_bp)

# Inicializar la base de datos al arrancar
init_db()


if __name__ == '__main__':
    debug_mode = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    app.run(debug=debug_mode)



  # { "workspaceRoot": "file:///vsls:/", "fileUri": "file:///vsls:/app.py" }