from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def pdf_safe(value):
    if value is None:
        return ''
    text = str(value)
    replacements = {
        '\u20ac': 'EUR', '\u00a3': 'GBP', '\u00a5': 'JPY',
        '\u2014': '-', '\u2013': '-', '\u2022': '-',
        '\u2122': 'TM', '\u00a9': '(c)', '\u00ae': '(R)',
        '\u2026': '...', '\u00b2': '2', '\u00b3': '3', '\u00b0': 'deg',
        '\u221a': 'sqrt', '\u00d7': 'x', '\u00f7': '/',
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    accents = {
        '\u00e1': 'a', '\u00e9': 'e', '\u00ed': 'i', '\u00f3': 'o', '\u00fa': 'u',
        '\u00e0': 'a', '\u00e8': 'e', '\u00ec': 'i', '\u00f2': 'o', '\u00f9': 'u',
        '\u00e4': 'a', '\u00eb': 'e', '\u00ef': 'i', '\u00f6': 'o', '\u00fc': 'u',
        '\u00e3': 'a', '\u00f5': 'o', '\u00f1': 'n',
        '\u00c1': 'A', '\u00c9': 'E', '\u00cd': 'I', '\u00d3': 'O', '\u00da': 'U',
        '\u00c0': 'A', '\u00c8': 'E', '\u00cc': 'I', '\u00d2': 'O', '\u00d9': 'U',
        '\u00c4': 'A', '\u00cb': 'E', '\u00cf': 'I', '\u00d6': 'O', '\u00dc': 'U',
        '\u00c3': 'A', '\u00d5': 'O', '\u00d1': 'N',
        '\u00e7': 'c', '\u00c7': 'C', '\u00df': 'ss',
    }
    for old, new in accents.items():
        text = text.replace(old, new)
    return ''.join(c if ord(c) < 128 else '?' for c in text)


def pdf_val(value, default='-'):
    text = pdf_safe(value)
    return text if text else default


def _style_header_row(ws, col_count):
    header_font = Font(name='Manrope', bold=True, size=10, color='FFFFFF')
    header_fill = PatternFill(start_color='000410', end_color='000410', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='D4BC9A'),
        right=Side(style='thin', color='D4BC9A'),
        top=Side(style='thin', color='D4BC9A'),
        bottom=Side(style='thin', color='D4BC9A'),
    )
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def _apply_data_border(ws, row, col_count):
    border = Border(
        left=Side(style='thin', color='E8D5B7'),
        right=Side(style='thin', color='E8D5B7'),
        top=Side(style='thin', color='E8D5B7'),
        bottom=Side(style='thin', color='E8D5B7'),
    )
    for col in range(1, col_count + 1):
        ws.cell(row=row, column=col).border = border
