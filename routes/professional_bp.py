import csv
import io
import logging
import os
import re

from datetime import datetime
from io import StringIO

logger = logging.getLogger(__name__)

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
import pytz

from flask import Blueprint, flash, jsonify, redirect, render_template, request, Response, send_file, session, url_for
from fpdf import FPDF
from werkzeug.utils import secure_filename

import config
import models
import rate_limit
import utils
from decorators import login_required, professional_required
from i18n import t, get_language
from services.database import date_format_sql, now_sql
from services.pdf_helpers import pdf_safe, pdf_val, _style_header_row, _apply_data_border
from utils import allowed_file, convert_to_argentina_time

professional_bp = Blueprint('professional', __name__, url_prefix='')


@professional_bp.route('/profesional')
@professional_required
def professional_view():
    conn = None
    try:
        conn = models.get_db_connection()
        user = conn.execute('SELECT username, doc_path FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        if not user:
            return redirect(url_for('public.index'))

        professional = conn.execute('SELECT status FROM professionals WHERE name = ?', (user['username'],)).fetchone()
        if not professional or professional['status'] != 'approved':
            return render_template('professional.html', pending=True, doc_path=user['doc_path'])

        return render_template('professional.html', pending=False, doc_path=user['doc_path'])
    finally:
        if conn:
            conn.close()


@professional_bp.route('/profesional/lead/<int:lead_id>')
@professional_required
def lead_detail(lead_id):
    conn = None
    try:
        conn = models.get_db_connection()
        user = conn.execute('SELECT username FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        if not user:
            return redirect(url_for('public.index'))

        professional = conn.execute('SELECT status FROM professionals WHERE name = ?', (user['username'],)).fetchone()
        if not professional or professional['status'] != 'approved':
            return render_template('professional.html', pending=True)

        lead = conn.execute('SELECT * FROM leads WHERE id = ?', (lead_id,)).fetchone()
    finally:
        if conn:
            conn.close()

    if not lead:
        return redirect(url_for('professional.professional_view'))

    lead_dict = dict(lead)
    lead_dict['timestamp'] = convert_to_argentina_time(lead_dict['timestamp'])
    phone_raw = lead_dict.get('phone') or ''
    lead_dict['phone_e164'] = utils.normalize_phone_to_e164(phone_raw)
    lead_dict['phone_is_mobile'] = bool(lead_dict['phone_e164'] and utils.is_whatsapp_capable(lead_dict['phone_e164']))
    return render_template('lead_detail.html', lead=lead_dict)


@professional_bp.route('/api/leads')
@professional_required
def get_leads_api():
    conn = None
    try:
        conn = models.get_db_connection()
        lang = get_language()

        user = conn.execute('SELECT username FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        if not user:
            return jsonify({"status": "error", "message": t('prof.access_denied', lang)}), 403

        professional = conn.execute('SELECT status FROM professionals WHERE name = ?', (user['username'],)).fetchone()
        if not professional or professional['status'] != 'approved':
            return jsonify({"status": "error", "message": t('prof.account_pending', lang)}), 403

        my_leads = request.args.get('my_leads', '1').strip() == '1'
        search = request.args.get('search', '').strip()
        type_filter = request.args.get('type', '').strip()
        prop_type = request.args.get('property_type', '').strip()
        zone_filter = request.args.get('zone', '').strip()
        min_budget = request.args.get('min_budget', '').strip()
        max_budget = request.args.get('max_budget', '').strip()
        budget_range = request.args.get('budget_range', '').strip()
        currency_filter = request.args.get('currency', '').strip()
        sort_by = request.args.get('sort', 'timestamp')
        sort_order = request.args.get('order', 'desc')
        time_range = request.args.get('time_range', '').strip()
        page = max(1, int(request.args.get('page', '1')))
        per_page = max(5, min(100, int(request.args.get('per_page', '25'))))
        offset = (page - 1) * per_page

        BUDGET_RANGES = {
            'hasta_200k': (0, 200000),
            '200k_500k': (200000, 500000),
            '500k_1m': (500000, 1000000),
            '1m_2m': (1000000, 2000000),
            'mas_2m': (2000000, None),
        }
        if budget_range and budget_range in BUDGET_RANGES:
            rng = BUDGET_RANGES[budget_range]
            min_budget = str(rng[0]) if rng[0] else ''
            max_budget = str(rng[1]) if rng[1] else ''

        query = 'SELECT * FROM leads WHERE 1=1'
        params = []

        if my_leads:
            conds, geo_params = _get_pro_geo_filter(conn, session['user_id'])
            for c in conds:
                query += f' AND {c}'
            params.extend(geo_params)

        if search:
            query += ' AND (zone LIKE ? OR email LIKE ? OR type LIKE ? OR budget LIKE ?)'
            search_param = f'%{search}%'
            params.extend([search_param, search_param, search_param, search_param])

        if type_filter:
            query += ' AND type = ?'
            params.append(type_filter)

        if prop_type:
            query += ' AND property_type = ?'
            params.append(prop_type)

        if zone_filter:
            query += ' AND zone LIKE ?'
            params.append(f'%{zone_filter}%')

        if min_budget:
            try:
                min_val = float(min_budget)
                query += " AND CAST(REPLACE(REPLACE(budget, '.', ''), ',', '') AS REAL) >= ?"
                params.append(min_val)
            except ValueError:
                pass

        if max_budget:
            try:
                max_val = float(max_budget)
                query += " AND CAST(REPLACE(REPLACE(budget, '.', ''), ',', '') AS REAL) <= ?"
                params.append(max_val)
            except ValueError:
                pass

        if currency_filter:
            query += ' AND currency = ?'
            params.append(currency_filter)

        valid_sort_fields = ['id', 'type', 'zone', 'budget', 'timestamp', 'email']
        if sort_by not in valid_sort_fields:
            sort_by = 'timestamp'

        order = 'DESC' if sort_order.lower() == 'desc' else 'ASC'
        if order not in ('ASC', 'DESC'):
            order = 'ASC'

        if time_range:
            try:
                tr = int(time_range)
                if tr > 0:
                    query += " AND timestamp >= datetime('now', ?)"
                    params.append(f'-{tr} days')
            except ValueError:
                pass

        count_query = f'SELECT COUNT(*) as total FROM ({query})'
        total = conn.execute(count_query, params).fetchone()['total']

        query += f' ORDER BY {sort_by} {order}, id DESC LIMIT ? OFFSET ?'
        params.extend([per_page, offset])

        leads = conn.execute(query, params).fetchall()

        leads_list = []
        for lead in leads:
            lead_dict = dict(lead)
            lead_dict['timestamp'] = convert_to_argentina_time(lead_dict['timestamp'])
            phone_raw = lead_dict.get('phone') or ''
            phone_e164 = utils.normalize_phone_to_e164(phone_raw)
            lead_dict['phone_e164'] = phone_e164
            lead_dict['phone_is_mobile'] = bool(phone_e164 and utils.is_whatsapp_capable(phone_e164))
            leads_list.append(lead_dict)

        professional_id = session['user_id']
        lead_ids = [lead['id'] for lead in leads_list]

        tracking_map = {}
        if lead_ids:
            placeholders = ','.join(['?'] * len(lead_ids))
            tracking_rows = conn.execute(
                f'SELECT lead_id, seen, contacted FROM lead_tracking WHERE professional_id = ? AND lead_id IN ({placeholders})',
                [professional_id] + lead_ids
            ).fetchall()
            for row in tracking_rows:
                tracking_map[row['lead_id']] = {
                    'seen': bool(row['seen']),
                    'contacted': bool(row['contacted'])
                }

        for lead in leads_list:
            tracking = tracking_map.get(lead['id'], {'seen': False, 'contacted': False})
            lead['tracking'] = tracking

        total_pages = max(1, (total + per_page - 1) // per_page)

        return jsonify({
            "success": True,
            "leads": leads_list,
            "total": total,
            "page": page,
            "per_page": per_page,
            "total_pages": total_pages
        })
    except Exception as e:
        logger.exception('Error en get_leads_api')
        return jsonify({"status": "error", "message": t('prof.internal_error', lang)}), 500
    finally:
        if conn:
            conn.close()


@professional_bp.route('/api/leads/filter-options')
@professional_required
def get_leads_filter_options():
    lang = get_language()
    conn = None
    try:
        conn = models.get_db_connection()
        user = conn.execute('SELECT username FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        if user:
            professional = conn.execute('SELECT status FROM professionals WHERE name = ?', (user['username'],)).fetchone()
            if not professional or professional['status'] != 'approved':
                return jsonify({"error": t('prof.approval_required', lang)}), 403
    finally:
        if conn:
            conn.close()

    from app_setup import filter_cache
    cached = filter_cache.get('filter_options')
    if cached:
        return jsonify(cached)

    conn = None
    try:
        conn = models.get_db_connection()
        types = [r[0] for r in conn.execute('SELECT DISTINCT type FROM leads WHERE type IS NOT NULL ORDER BY type').fetchall()]
        prop_types = [r[0] for r in conn.execute('SELECT DISTINCT property_type FROM leads WHERE property_type IS NOT NULL ORDER BY property_type').fetchall()]
        currencies = [r[0] for r in conn.execute('SELECT DISTINCT currency FROM leads WHERE currency IS NOT NULL ORDER BY currency').fetchall()]
        zones = [r[0] for r in conn.execute('SELECT DISTINCT zone FROM leads WHERE zone IS NOT NULL ORDER BY zone').fetchall()]

        result = {
            'types': types,
            'property_types': prop_types,
            'currencies': currencies,
            'zones': zones,
        }
        filter_cache.set('filter_options', result)
        return jsonify(result)
    finally:
        if conn:
            conn.close()


def _get_pro_geo_filter(conn, user_id):
    pro_data = conn.execute(
        'SELECT province, zone FROM professionals WHERE user_id = ?',
        (user_id,)
    ).fetchone()
    conditions = []
    params = []
    if pro_data:
        pro_province = (pro_data['province'] or '').strip()
        pro_zone = (pro_data['zone'] or '').strip()
        if pro_province:
            conditions.append('province = ?')
            params.append(pro_province)
        if pro_zone:
            conditions.append('zone LIKE ?')
            params.append(f'%{pro_zone}%')
    return conditions, params


def _query_leads_stats(conn, user_id, my_leads, month):
    """Helper: build stats data dict. conn must be open, caller owns close."""
    where_clauses = ['1=1']
    params = []

    if my_leads:
        conds, geo_params = _get_pro_geo_filter(conn, user_id)
        where_clauses.extend(conds)
        params.extend(geo_params)

    where_sql = ' AND '.join(where_clauses)

    cur_params = params + [month]
    cur_stats = conn.execute(f'''
        SELECT
            COUNT(*) as total,
            AVG(CAST(SUBSTR(budget, 1, INSTR(budget, ' ') - 1) AS REAL)) as avg_budget
        FROM leads WHERE {where_sql} AND {date_format_sql('timestamp', '%Y-%m')} = ?
    ''', cur_params).fetchone()

    cur_year, cur_mon = month.split('-')
    mon_int = int(cur_mon)
    prev_year = str(int(cur_year) - 1) if mon_int == 1 else cur_year
    prev_mon = f'12' if mon_int == 1 else f'{mon_int - 1:02d}'
    prev_month = f'{prev_year}-{prev_mon}'
    prev_params = params + [prev_month]
    prev_stats = conn.execute(f'''
        SELECT COUNT(*) as total FROM leads
        WHERE {where_sql} AND {date_format_sql('timestamp', '%Y-%m')} = ?
    ''', prev_params).fetchone()

    pt_params = params + [month]
    by_property_type = conn.execute(f'''
        SELECT property_type, COUNT(*) as count
        FROM leads WHERE {where_sql} AND {date_format_sql('timestamp', '%Y-%m')} = ?
        GROUP BY property_type ORDER BY count DESC
    ''', pt_params).fetchall()

    z_params = params + [month]
    by_zone = conn.execute(f'''
        SELECT zone, COUNT(*) as count
        FROM leads WHERE {where_sql} AND {date_format_sql('timestamp', '%Y-%m')} = ?
        GROUP BY zone ORDER BY count DESC LIMIT 10
    ''', z_params).fetchall()

    ot_params = params + [month]
    by_operation_type = conn.execute(f'''
        SELECT type, COUNT(*) as count
        FROM leads WHERE {where_sql} AND {date_format_sql('timestamp', '%Y-%m')} = ?
        GROUP BY type ORDER BY count DESC
    ''', ot_params).fetchall()

    trend = conn.execute(f'''
        SELECT {date_format_sql('timestamp', '%Y-%m')} as month, COUNT(*) as count
        FROM leads WHERE {where_sql}
        GROUP BY month ORDER BY month DESC LIMIT 6
    ''', params).fetchall()

    active_zone_count = conn.execute(f'''
        SELECT COUNT(DISTINCT zone) as cnt
        FROM leads WHERE {where_sql} AND {date_format_sql('timestamp', '%Y-%m')} = ? AND zone != ''
    ''', params + [month]).fetchone()[0]

    return {
        'total': cur_stats['total'] or 0,
        'avg_budget': round(cur_stats['avg_budget'] or 0, 0),
        'by_property_type': [{'label': r['property_type'], 'value': r['count']} for r in by_property_type],
        'by_zone': [{'label': r['zone'], 'value': r['count']} for r in by_zone],
        'by_operation_type': [{'label': r['type'], 'value': r['count']} for r in by_operation_type],
        'trend': [{'label': r['month'], 'value': r['count']} for r in reversed(trend)],
        'previous_month': {'total': prev_stats['total'] or 0},
        'active_zones': active_zone_count,
        'month': month,
    }


@professional_bp.route('/api/leads/stats')
@professional_required
def get_leads_stats():
    conn = None
    try:
        conn = models.get_db_connection()
        lang = get_language()
        user = conn.execute('SELECT username FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        if not user:
            return jsonify({'status': 'error', 'message': t('prof.access_denied', lang)}), 403

        professional = conn.execute('SELECT status FROM professionals WHERE name = ?', (user['username'],)).fetchone()
        if not professional or professional['status'] != 'approved':
            return jsonify({'status': 'error', 'message': t('prof.account_pending', lang)}), 403

        my_leads = request.args.get('my_leads', '1').strip() == '1'
        month = request.args.get('month', '').strip()
        if not month or not re.match(r'^\d{4}-\d{2}$', month):
            month = datetime.now().strftime('%Y-%m')

        stats = _query_leads_stats(conn, session['user_id'], my_leads, month)
        return jsonify({'status': 'success', 'stats': stats, 'month': month})
    except Exception as e:
        logger.exception('Error en get_leads_stats')
        return jsonify({'status': 'error', 'message': t('prof.internal_error', lang)}), 500
    finally:
        if conn:
            conn.close()


@professional_bp.route('/api/leads/stats/export')
@professional_required
def export_stats_csv():
    conn = None
    try:
        conn = models.get_db_connection()
        lang = get_language()
        user = conn.execute('SELECT username FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        if not user:
            return t('prof.access_denied', lang), 403

        professional = conn.execute('SELECT status FROM professionals WHERE name = ?', (user['username'],)).fetchone()
        if not professional or professional['status'] != 'approved':
            return t('prof.account_pending', lang), 403

        my_leads = request.args.get('my_leads', '1').strip() == '1'
        month = request.args.get('month', '').strip()
        if not month or not re.match(r'^\d{4}-\d{2}$', month):
            month = datetime.now().strftime('%Y-%m')

        stats = _query_leads_stats(conn, session['user_id'], my_leads, month)

        def generate():
            data = StringIO()
            writer = csv.writer(data)

            # Summary section
            writer.writerow([t('prof.export_summary', lang)])
            writer.writerow([t('prof.export_month', lang), stats.get('month', month)])
            writer.writerow(['Total Leads', stats['total']])
            writer.writerow([t('prof.export_avg_budget', lang), f"${stats['avg_budget']:,.0f}"])
            writer.writerow([t('prof.export_vs_prev_month', lang), stats['previous_month']['total']])
            writer.writerow([t('prof.export_active_zones', lang), stats['active_zones']])
            writer.writerow([])

            # Property type section
            writer.writerow([t('prof.export_property_type_section', lang)])
            writer.writerow([t('prof.export_type', lang), t('prof.export_count', lang)])
            for pt in stats['by_property_type']:
                writer.writerow([pt['label'], pt['value']])
            writer.writerow([])

            # Zone section
            writer.writerow([t('prof.export_zones_section', lang)])
            writer.writerow([t('prof.export_zone', lang), t('prof.export_count', lang)])
            for z in stats['by_zone']:
                writer.writerow([z['label'], z['value']])
            writer.writerow([])

            # Operation type section
            writer.writerow([t('prof.export_operation_type_section', lang)])
            writer.writerow([t('prof.export_operation', lang), t('prof.export_count', lang)])
            for ot in stats['by_operation_type']:
                writer.writerow([ot['label'], ot['value']])
            writer.writerow([])

            # Trend section
            writer.writerow([t('prof.export_trend_section', lang)])
            writer.writerow([t('prof.export_month', lang), 'Leads'])
            for trend in stats['trend']:
                writer.writerow([trend['label'], trend['value']])

            yield data.getvalue()
            data.seek(0)
            data.truncate(0)

        filename = f'estadisticas_{month}_{datetime.now().strftime("%Y%m%d_%H%M")}.csv'
        return Response(
            generate(),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )
    except Exception as e:
        logger.exception('Error en export_stats_csv')
        return t('prof.internal_error', lang), 500
    finally:
        if conn:
            conn.close()


@professional_bp.route('/api/leads/stats/export/xlsx')
@professional_required
def export_stats_xlsx():
    conn = None
    try:
        conn = models.get_db_connection()
        lang = get_language()
        user = conn.execute('SELECT username FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        if not user:
            return t('prof.access_denied', lang), 403

        professional = conn.execute('SELECT status FROM professionals WHERE name = ?', (user['username'],)).fetchone()
        if not professional or professional['status'] != 'approved':
            return t('prof.account_pending', lang), 403

        my_leads = request.args.get('my_leads', '1').strip() == '1'
        month = request.args.get('month', '').strip()
        if not month or not re.match(r'^\d{4}-\d{2}$', month):
            month = datetime.now().strftime('%Y-%m')

        stats = _query_leads_stats(conn, session['user_id'], my_leads, month)

        wb = openpyxl.Workbook()

        # Shared data font
        data_font = Font(name='Manrope', size=10, color='000410')
        label_font = Font(name='Manrope', size=10, bold=True, color='735A3A')
        title_font = Font(name='Manrope', size=10, bold=True, color='000410')

        # ─── Sheet 1: Resumen ───
        ws_resumen = wb.active
        ws_resumen.title = t('prof.export_summary_sheet', lang)
        resumen_data = [
            [t('prof.export_month', lang), stats.get('month', month)],
            ['Total Leads', stats['total']],
            [t('prof.export_avg_budget', lang), stats['avg_budget']],
            [t('prof.export_vs_prev_month', lang), stats['previous_month']['total']],
            [t('prof.export_active_zones', lang), stats['active_zones']],
        ]
        for col, h in enumerate([t('prof.export_metric', lang), t('prof.export_value', lang)], 1):
            cell = ws_resumen.cell(row=1, column=col, value=h)
        _style_header_row(ws_resumen, 2)
        ws_resumen.column_dimensions['A'].width = 28
        ws_resumen.column_dimensions['B'].width = 18

        for row_idx, row_data in enumerate(resumen_data, 2):
            for col, val in enumerate(row_data, 1):
                cell = ws_resumen.cell(row=row_idx, column=col, value=val)
                cell.font = data_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if col == 1:
                    cell.font = label_font
            _apply_data_border(ws_resumen, row_idx, 2)

        # Separator + note rows
        note_row = len(resumen_data) + 3
        ws_resumen.cell(row=note_row, column=1, value=t('prof.export_generated_by', lang)).font = Font(
            name='Manrope', size=8, italic=True, color='A68A64'
        )

        # ─── Sheet 2: Tipo de Propiedad ───
        ws_pt = wb.create_sheet(t('prof.export_property_type_sheet', lang))
        for col, h in enumerate([t('prof.export_type', lang), t('prof.export_count', lang)], 1):
            ws_pt.cell(row=1, column=col, value=h)
        _style_header_row(ws_pt, 2)
        ws_pt.column_dimensions['A'].width = 32
        ws_pt.column_dimensions['B'].width = 14
        ws_pt.auto_filter.ref = f'A1:B{len(stats["by_property_type"]) + 1}'
        for row_idx, pt in enumerate(stats['by_property_type'], 2):
            c1 = ws_pt.cell(row=row_idx, column=1, value=pt['label'])
            c1.font = label_font
            c1.alignment = Alignment(horizontal='left', vertical='center')
            c2 = ws_pt.cell(row=row_idx, column=2, value=pt['value'])
            c2.font = data_font
            c2.alignment = Alignment(horizontal='center', vertical='center')
            _apply_data_border(ws_pt, row_idx, 2)

        # ─── Sheet 3: Zonas ───
        ws_z = wb.create_sheet(t('prof.export_zones_sheet', lang))
        for col, h in enumerate([t('prof.export_zone', lang), t('prof.export_count', lang)], 1):
            ws_z.cell(row=1, column=col, value=h)
        _style_header_row(ws_z, 2)
        ws_z.column_dimensions['A'].width = 34
        ws_z.column_dimensions['B'].width = 14
        ws_z.auto_filter.ref = f'A1:B{len(stats["by_zone"]) + 1}'
        for row_idx, z in enumerate(stats['by_zone'], 2):
            c1 = ws_z.cell(row=row_idx, column=1, value=z['label'])
            c1.font = label_font
            c1.alignment = Alignment(horizontal='left', vertical='center')
            c2 = ws_z.cell(row=row_idx, column=2, value=z['value'])
            c2.font = data_font
            c2.alignment = Alignment(horizontal='center', vertical='center')
            _apply_data_border(ws_z, row_idx, 2)

        # ─── Sheet 4: Tipo de Operación ───
        ws_ot = wb.create_sheet(t('prof.export_operation_type_sheet', lang))
        for col, h in enumerate([t('prof.export_operation', lang), t('prof.export_count', lang)], 1):
            ws_ot.cell(row=1, column=col, value=h)
        _style_header_row(ws_ot, 2)
        ws_ot.column_dimensions['A'].width = 34
        ws_ot.column_dimensions['B'].width = 14
        ws_ot.auto_filter.ref = f'A1:B{len(stats["by_operation_type"]) + 1}'
        for row_idx, ot in enumerate(stats['by_operation_type'], 2):
            c1 = ws_ot.cell(row=row_idx, column=1, value=ot['label'])
            c1.font = label_font
            c1.alignment = Alignment(horizontal='left', vertical='center')
            c2 = ws_ot.cell(row=row_idx, column=2, value=ot['value'])
            c2.font = data_font
            c2.alignment = Alignment(horizontal='center', vertical='center')
            _apply_data_border(ws_ot, row_idx, 2)

        # ─── Sheet 5: Tendencia Mensual ───
        ws_t = wb.create_sheet(t('prof.export_trend_sheet', lang))
        for col, h in enumerate([t('prof.export_month', lang), 'Leads'], 1):
            ws_t.cell(row=1, column=col, value=h)
        _style_header_row(ws_t, 2)
        ws_t.column_dimensions['A'].width = 28
        ws_t.column_dimensions['B'].width = 14
        ws_t.auto_filter.ref = f'A1:B{len(stats["trend"]) + 1}'
        for row_idx, trend in enumerate(stats['trend'], 2):
            c1 = ws_t.cell(row=row_idx, column=1, value=trend['label'])
            c1.font = label_font
            c1.alignment = Alignment(horizontal='left', vertical='center')
            c2 = ws_t.cell(row=row_idx, column=2, value=trend['value'])
            c2.font = data_font
            c2.alignment = Alignment(horizontal='center', vertical='center')
            _apply_data_border(ws_t, row_idx, 2)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f'estadisticas_{month}_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.exception('Error en export_stats_xlsx')
        return t('prof.internal_error', lang), 500
    finally:
        if conn:
            conn.close()


@professional_bp.route('/api/leads/filter-options/invalidate', methods=['POST'])
@professional_required
def invalidate_filter_cache():
    from app_setup import filter_cache
    lang = get_language()
    filter_cache.invalidate()
    return jsonify({"status": "success", "message": t('prof.cache_invalidated', lang)})


@professional_bp.route('/api/leads/export')
@professional_required
def export_leads_csv():
    conn = None
    leads = []
    try:
        conn = models.get_db_connection()
        lang = get_language()
        user = conn.execute('SELECT username FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        if not user:
            return t('prof.access_denied', lang), 403

        professional = conn.execute('SELECT status FROM professionals WHERE name = ?', (user['username'],)).fetchone()
        if not professional or professional['status'] != 'approved':
            return t('prof.account_pending', lang), 403

        conds, geo_params = _get_pro_geo_filter(conn, session['user_id'])
        query = 'SELECT id, type, zone, budget, currency, timestamp FROM leads WHERE 1=1'
        params = []
        for c in conds:
            query += f' AND {c}'
        params.extend(geo_params)
        query += ' ORDER BY timestamp DESC'
        leads = conn.execute(query, params).fetchall()
    finally:
        if conn:
            conn.close()

    def generate():
        data = StringIO()
        writer = csv.writer(data)
        writer.writerow(t('prof.export_csv_headers', lang))
        yield data.getvalue()
        data.seek(0)
        data.truncate(0)

        for lead in leads:
            timestamp_argentina = convert_to_argentina_time(lead['timestamp'])
            writer.writerow([lead['id'], lead['type'], lead['zone'], lead['budget'], lead['currency'], timestamp_argentina])
            yield data.getvalue()
            data.seek(0)
            data.truncate(0)

    filename = f"leads_archestate_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return Response(
        generate(),
        mimetype='text/csv',
        headers={
            'Content-Disposition': f'attachment; filename="{filename}"'
        }
    )


@professional_bp.route('/api/leads/export/xlsx')
@professional_required
def export_leads_xlsx():
    conn = None
    leads = []
    try:
        conn = models.get_db_connection()
        lang = get_language()
        user = conn.execute('SELECT username FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        if not user:
            return t('prof.access_denied', lang), 403

        professional = conn.execute('SELECT status FROM professionals WHERE name = ?', (user['username'],)).fetchone()
        if not professional or professional['status'] != 'approved':
            return t('prof.account_pending', lang), 403

        conds, geo_params = _get_pro_geo_filter(conn, session['user_id'])
        query = 'SELECT id, type, zone, budget, currency, timestamp FROM leads WHERE 1=1'
        params = []
        for c in conds:
            query += f' AND {c}'
        params.extend(geo_params)
        query += ' ORDER BY timestamp DESC'
        leads = conn.execute(query, params).fetchall()
    finally:
        if conn:
            conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = t('prof.export_leads_sheet', lang)

    headers = t('prof.export_leads_headers', lang)
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    _style_header_row(ws, len(headers))

    col_widths = [10, 30, 28, 18, 12, 24]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    data_font = Font(name='Manrope', size=10, color='000410')
    data_align = Alignment(horizontal='center', vertical='center')
    date_font = Font(name='Manrope', size=10, color='A68A64')

    for row_num, lead in enumerate(leads, 2):
        timestamp_argentina = convert_to_argentina_time(lead['timestamp'])
        vals = [lead['id'], lead['type'], lead['zone'], lead['budget'], lead['currency'], timestamp_argentina]
        for col_num, val in enumerate(vals, 1):
            cell = ws.cell(row=row_num, column=col_num, value=val)
            cell.font = data_font
            cell.alignment = data_align
            if col_num == len(vals):
                cell.font = date_font
        _apply_data_border(ws, row_num, len(headers))

    ws.auto_filter.ref = f'A1:F{len(leads) + 1}'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"leads_archestate_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@professional_bp.route('/api/lead/<int:lead_id>/download')
@professional_required
def download_lead_pdf(lead_id):
    conn = None
    try:
        conn = models.get_db_connection()
        lang = get_language()
        user = conn.execute('SELECT username FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        if not user:
            return t('prof.access_denied', lang), 403

        professional = conn.execute('SELECT status FROM professionals WHERE name = ?', (user['username'],)).fetchone()
        if not professional or professional['status'] != 'approved':
            return t('prof.account_pending', lang), 403

        lead = conn.execute('SELECT * FROM leads WHERE id = ?', (lead_id,)).fetchone()
    finally:
        if conn:
            conn.close()

    if not lead:
        return jsonify({"status": "error", "message": t('prof.lead_not_found', lang)}), 404

    lead = dict(lead)

    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    midnight = (0, 4, 16)
    gold = (115, 90, 58)

    pdf.set_font('Times', 'BI', 20)
    pdf.set_text_color(*midnight)
    pdf.cell(0, 15, t('prof.pdf_title', lang), ln=True, align='C')
    pdf.ln(5)

    pdf.set_font('Helvetica', '', 10)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 8, t('prof.pdf_subtitle', lang, lead_id=lead['id']), ln=True, align='C')
    pdf.ln(10)

    def section_header(title):
        pdf.set_fill_color(*gold)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font('Helvetica', 'B', 12)
        pdf.cell(0, 8, title.upper(), ln=True, fill=True)
        pdf.set_text_color(*midnight)
        pdf.set_font('Helvetica', '', 10)
        pdf.ln(2)

    section_header(t('prof.pdf_operation_type', lang))
    pdf.cell(0, 6, pdf_val(lead['type']), ln=True)

    prop_type = pdf_safe(lead.get('property_type', '')).lower()
    prop_type_label = {'departamento': 'Departamento', 'casa': 'Casa', 'duplex': 'Duplex',
                       'penthouse': 'Penthouse', 'local_comercial': 'Local Comercial'}.get(prop_type, prop_type)
    section_header(t('prof.pdf_property_type', lang))
    pdf.cell(0, 6, pdf_val(prop_type_label, t('prof.not_specified', lang)), ln=True)

    section_header(t('prof.pdf_geographic_zone', lang))
    pdf.cell(0, 6, f"{t('prof.pdf_zone_prefix', lang)}{pdf_val(lead['zone'])}", ln=True)
    pdf.cell(0, 6, f"{t('prof.pdf_province_prefix', lang)}{pdf_val(lead.get('province'), t('prof.not_specified_fem', lang))}", ln=True)

    section_header(t('prof.pdf_budget', lang))
    budget_symbol = 'USD' if lead['currency'] == 'USD' else 'EUR' if lead['currency'] == 'EUR' else '$'
    pdf.cell(0, 6, f"{budget_symbol} {pdf_val(lead['budget'])}", ln=True)

    section_header(t('prof.pdf_architectural_style', lang))
    pdf.cell(0, 6, pdf_val(lead.get('architectural_style'), t('prof.not_specified', lang)), ln=True)

    section_header(t('prof.pdf_direct_contact', lang))
    pdf.cell(0, 6, f"Email: {pdf_val(lead['email'])}", ln=True)
    pdf.cell(0, 6, f"Telefono: {pdf_val(lead['phone'])}", ln=True)

    section_header(t('prof.pdf_registered', lang))
    pdf.cell(0, 6, pdf_val(convert_to_argentina_time(lead['timestamp'])), ln=True)
    pdf.ln(5)

    section_header(t('prof.pdf_tech_specs', lang))

    pdf.set_font('Helvetica', 'B', 10)
    pdf.cell(60, 8, t('prof.pdf_rooms', lang), border=1)
    pdf.cell(0, 8, pdf_val(lead.get('ambientes')), ln=True, border=1)

    pdf.cell(60, 8, t('prof.pdf_bedrooms', lang), border=1)
    pdf.cell(0, 8, pdf_val(lead['bedrooms']), ln=True, border=1)

    pdf.cell(60, 8, t('prof.pdf_bathrooms', lang), border=1)
    pdf.cell(0, 8, pdf_val(lead['bathrooms']), ln=True, border=1)

    pdf.cell(60, 8, t('prof.pdf_parking', lang), border=1)
    pdf.cell(0, 8, pdf_val(lead.get('parking'), t('prof.not_specified', lang)), ln=True, border=1)

    pdf.cell(60, 8, t('prof.pdf_orientation', lang), border=1)
    pdf.cell(0, 8, pdf_val(lead.get('orientation'), t('prof.not_specified_fem', lang)), ln=True, border=1)

    pdf.ln(2)

    pdf.cell(60, 8, t('prof.pdf_usable_m2', lang), border=1)
    pdf.cell(0, 8, f"{pdf_val(lead.get('usable_m2'))} m2" if lead.get('usable_m2') else '-', ln=True, border=1)

    pdf.cell(60, 8, t('prof.pdf_total_area', lang), border=1)
    pdf.cell(0, 8, f"{pdf_val(lead.get('total_area'))} m2" if lead.get('total_area') else '-', ln=True, border=1)

    pdf.cell(60, 8, t('prof.pdf_land_area', lang), border=1)
    pdf.cell(0, 8, f"{pdf_val(lead.get('land_area'))} m2" if lead.get('land_area') else '-', ln=True, border=1)

    pdf.cell(60, 8, t('prof.pdf_built_area', lang), border=1)
    pdf.cell(0, 8, f"{pdf_val(lead.get('built_area'))} m2" if lead.get('built_area') else '-', ln=True, border=1)

    pdf.ln(5)

    section_header(t('prof.pdf_condition_age', lang))
    pdf.set_font('Helvetica', 'B', 10)
    pdf.cell(60, 8, t('prof.pdf_condition', lang), border=1)
    pdf.cell(0, 8, pdf_val(lead.get('property_condition'), t('prof.not_specified', lang)), ln=True, border=1)

    pdf.cell(60, 8, t('prof.pdf_age', lang), border=1)
    pdf.cell(0, 8, pdf_val(lead.get('property_age'), t('prof.not_specified_fem', lang)), ln=True, border=1)

    pdf.ln(5)

    section_header(t('prof.pdf_amenities', lang))
    amenities = lead.get('amenities', '')
    if amenities and str(amenities).strip():
        for amenity in pdf_safe(amenities).split(','):
            stripped = amenity.strip()
            if stripped:
                pdf.cell(0, 6, f"- {stripped}", ln=True)
    else:
        pdf.cell(0, 6, t('prof.not_specified_fem_pl', lang), ln=True)

    section_header(t('prof.pdf_property_details', lang))
    if prop_type == 'departamento':
        pdf.cell(0, 6, f"{t('prof.pdf_floor_block', lang)}{pdf_val(lead.get('floor_block'), t('prof.not_specified', lang))}", ln=True)
    pdf.cell(0, 6, f"{t('prof.pdf_pool', lang)}{pdf_val(lead.get('pool'), t('prof.not_specified_fem', lang))}", ln=True)
    pdf.cell(0, 6, f"{t('prof.pdf_elevator', lang)}{pdf_val(lead.get('elevator'), t('prof.not_specified', lang))}", ln=True)

    pdf_output = pdf.output(dest='S')
    if isinstance(pdf_output, str):
        pdf_output = pdf_output.encode('latin-1')

    buffer = io.BytesIO(pdf_output)
    buffer.seek(0)

    filename = f"lead_{lead['id']}.pdf"

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/pdf'
    )


@professional_bp.route('/api/lead/<int:lead_id>/toggle-status', methods=['POST'])
@professional_required
def toggle_lead_status(lead_id):
    conn = None
    try:
        conn = models.get_db_connection()
        lang = get_language()

        user = conn.execute('SELECT username FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        if not user:
            return jsonify({'error': t('prof.access_denied', lang)}), 403

        professional = conn.execute(
            'SELECT status FROM professionals WHERE name = ?',
            (user['username'],)
        ).fetchone()
        if not professional or professional['status'] != 'approved':
            return jsonify({'error': t('prof.account_pending', lang)}), 403

        lead = conn.execute('SELECT id FROM leads WHERE id = ?', (lead_id,)).fetchone()
        if not lead:
            return jsonify({'error': t('prof.lead_not_found', lang)}), 404

        data = request.get_json() or {}
        status_type = data.get('status')

        if status_type not in ('seen', 'contacted'):
            return jsonify({'error': t('prof.invalid_status_type', lang)}), 400

        professional_id = session['user_id']
        argentina_tz = pytz.timezone('America/Argentina/Buenos_Aires')
        now = datetime.now(argentina_tz).strftime('%Y-%m-%d %H:%M:%S')

        tracking = conn.execute(
            'SELECT * FROM lead_tracking WHERE professional_id = ? AND lead_id = ?',
            (professional_id, lead_id)
        ).fetchone()

        if tracking:
            current_value = tracking[status_type]
            new_value = 0 if current_value else 1
            timestamp_col = f'{status_type}_at'
            timestamp_value = now if new_value else None

            conn.execute(
                f'UPDATE lead_tracking SET {status_type} = ?, {timestamp_col} = ? WHERE professional_id = ? AND lead_id = ?',
                (new_value, timestamp_value, professional_id, lead_id)
            )
        else:
            seen_val = 1 if status_type == 'seen' else 0
            contacted_val = 1 if status_type == 'contacted' else 0
            seen_at = now if status_type == 'seen' else None
            contacted_at = now if status_type == 'contacted' else None

            conn.execute(
                'INSERT INTO lead_tracking (professional_id, lead_id, seen, contacted, seen_at, contacted_at) VALUES (?, ?, ?, ?, ?, ?)',
                (professional_id, lead_id, seen_val, contacted_val, seen_at, contacted_at)
            )
            new_value = 1

        conn.commit()

        if new_value:
            from services.notifications import notify_lead_status_change
            try:
                notify_lead_status_change(lead_id, professional_id, status_type)
            except Exception:
                pass

        return jsonify({
            'success': True,
            'status': status_type,
            'value': new_value,
            'timestamp': now if new_value else None
        })
    except Exception as e:
        logger.exception('Error en toggle_lead_status')
        return jsonify({'error': t('prof.internal_error_short', lang)}), 500
    finally:
        if conn:
            conn.close()


@professional_bp.route('/api/lead/<int:lead_id>/report', methods=['POST'])
@professional_required
def report_lead(lead_id):
    conn = None
    try:
        conn = models.get_db_connection()
        lang = get_language()

        user = conn.execute('SELECT username FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        if not user:
            return jsonify({'error': t('prof.access_denied', lang)}), 403

        professional = conn.execute(
            'SELECT status FROM professionals WHERE name = ?',
            (user['username'],)
        ).fetchone()
        if not professional or professional['status'] != 'approved':
            return jsonify({'error': t('prof.account_pending', lang)}), 403

        lead = conn.execute('SELECT id, type, phone FROM leads WHERE id = ?', (lead_id,)).fetchone()
        if not lead:
            return jsonify({'error': t('prof.lead_not_found', lang)}), 404

        data = request.get_json() or {}
        notes = utils.safe_text(data.get('notes', ''))[:500]

        existing = conn.execute(
            'SELECT id FROM lead_reports WHERE lead_id = ? AND reported_by = ? AND status = ?',
            (lead_id, session['user_id'], 'pending')
        ).fetchone()
        if existing:
            return jsonify({'error': t('prof.already_reported', lang)}), 400

        conn.execute(
            'INSERT INTO lead_reports (lead_id, reported_by, reason, notes, status) VALUES (?, ?, ?, ?, ?)',
            (lead_id, session['user_id'], 'telefono_inexistente', notes, 'pending')
        )
        conn.commit()

        utils.log_action(
            "Reporte de Lead",
            f"Lead ID: {lead_id} (phone_hash={utils.hash_phone_digits(lead['phone'] or '')}) reportado por {user['username']}",
            session
        )

        return jsonify({
            'success': True,
            'message': t('prof.report_submitted', lang)
        })
    except Exception as e:
        logger.exception('Error en report_lead')
        return jsonify({'error': t('prof.internal_error_short', lang)}), 500
    finally:
        if conn:
            conn.close()


@professional_bp.route('/api/professional/doc-status', methods=['GET'])
@professional_required
def get_doc_status():
    conn = None
    try:
        conn = models.get_db_connection()
        lang = get_language()
        user = conn.execute('SELECT doc_path FROM users WHERE id = ?', (session['user_id'],)).fetchone()

        if not user:
            return jsonify({"error": t('prof.user_not_found', lang)}), 404

        doc_path = user['doc_path']
        has_doc = bool(doc_path)

        if has_doc:
            from flask import current_app
            full_path = os.path.join(current_app.config['UPLOAD_FOLDER'], doc_path)
            has_doc = os.path.exists(full_path)

        return jsonify({
            "has_doc": has_doc,
            "filename": doc_path if has_doc else None,
            "display_name": re.sub(r'^user_\d+_', '', doc_path) if has_doc and doc_path else None,
        })
    except Exception as e:
        logger.exception('Error en get_doc_status')
        return jsonify({"error": t('prof.internal_error_short', lang)}), 500
    finally:
        if conn:
            conn.close()


MAX_UPLOAD_BYTES = 10 * 1024 * 1024


@professional_bp.route('/api/professional/upload', methods=['POST'])
@rate_limit.check_rate_limit(limit=100, window=60)
@professional_required
def upload_professional_doc():
    from flask import current_app
    lang = get_language()

    if 'document' not in request.files:
        return jsonify({"error": t('prof.no_file_included', lang)}), 400

    file = request.files['document']
    if not file or file.filename == '':
        return jsonify({"error": t('prof.no_file_selected', lang)}), 400

    if not allowed_file(file.filename):
        return jsonify({
            "error": t('prof.invalid_file_type', lang)
        }), 415

    mime_valid, detected_ext, mime_error = utils.validate_mime_type(file, file.filename)
    if not mime_valid:
        return jsonify({"error": mime_error}), 415

    file.seek(0, 2)
    size = file.tell()
    file.seek(0)
    if size > config.MAX_UPLOAD_SIZE:
        return jsonify({"error": t('prof.file_too_large', lang)}), 413

    original_name = secure_filename(file.filename)
    filename = f"user_{session['user_id']}_{original_name}"
    upload_dir = current_app.config['UPLOAD_FOLDER']

    os.makedirs(upload_dir, exist_ok=True)

    conn = None
    try:
        conn = models.get_db_connection()
        prev_user = conn.execute('SELECT doc_path FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        if prev_user and prev_user['doc_path']:
            prev_path = os.path.join(upload_dir, prev_user['doc_path'])
            if os.path.exists(prev_path):
                try:
                    os.remove(prev_path)
                except Exception:
                    pass

        file.save(os.path.join(upload_dir, filename))

        conn.execute('UPDATE users SET doc_path = ? WHERE id = ?', (filename, session['user_id']))
        conn.commit()
    finally:
        if conn:
            conn.close()

    utils.log_action("Subida de Documento", f"Usuario ID: {session['user_id']}", session)

    return jsonify({
        "status": "success",
        "message": t('prof.doc_uploaded', lang),
        "filename": filename,
        "display_name": original_name,
    })


@professional_bp.route('/profesional/download_doc')
@professional_required
def download_own_doc():
    from flask import current_app

    conn = None
    try:
        conn = models.get_db_connection()
        lang = get_language()
        user = conn.execute('SELECT doc_path FROM users WHERE id = ?', (session['user_id'],)).fetchone()

        if not user or not user['doc_path']:
            flash(t('prof.no_doc_yet', lang), 'error')
            return redirect(url_for('professional.professional_view'))

        directory = current_app.config['UPLOAD_FOLDER']
        filename = user['doc_path']

        if not os.path.exists(os.path.join(directory, filename)):
            flash(t('prof.file_not_found_server', lang, filename=filename), 'error')
            return redirect(url_for('professional.professional_view'))

        from flask import send_from_directory
        return send_from_directory(directory, filename, as_attachment=True)

    except Exception as e:
        logger.exception('Error en download_professional_doc')
        flash(t('prof.internal_server_error', lang), 'error')
        return redirect(url_for('professional.professional_view'))
    finally:
        if conn:
            conn.close()
