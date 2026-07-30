import os
import json
import logging
from datetime import datetime
from io import BytesIO

from flask import Blueprint, render_template, request, session, jsonify, redirect, url_for, current_app, send_file
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

import config
import models
import decorators
import validators
import utils
from utils import parse_budget
import rate_limit
from services.pdf_helpers import pdf_safe, pdf_val, _style_header_row, _apply_data_border
from i18n import t, get_language
from services.notifications import notify_admins


logger = logging.getLogger(__name__)


profile_bp = Blueprint('profile', __name__, url_prefix='')


ALLOWED_LEAD_EDIT_FIELDS = [
    'zone', 'province', 'budget', 'currency',
    'floor_block', 'usable_m2', 'elevator',
    'land_area', 'built_area', 'pool',
    'architectural_style', 'bedrooms', 'bathrooms',
    'total_area', 'amenities', 'ambientes',
    'parking', 'orientation', 'property_condition',
    'property_age', 'community_pool', 'additional_features'
]


@profile_bp.route('/mi-perfil')
@decorators.login_required
def profile_view():
    if session.get('role') == 'admin':
        return redirect(url_for('admin.admin_view'))
    user = models.get_user_profile(session['user_id'])
    return render_template('profile.html', user=user)


@profile_bp.route('/mi-perfil/lead/<int:lead_id>/editar')
@decorators.login_required
def edit_lead_view(lead_id):
    user_id = session['user_id']
    lead = models.get_lead_by_id_and_user(lead_id, user_id)
    if not lead:
        return redirect(url_for('profile.profile_view'))
    versions = models.get_lead_versions(lead_id)
    return render_template('edit_lead.html', lead=lead, versions=versions)


@profile_bp.route('/api/profile/leads', methods=['GET'])
@decorators.login_required
def api_get_user_leads():
    user_id = session['user_id']
    leads = models.get_user_leads(user_id)
    for lead in leads:
        if lead.get('timestamp'):
            lead['timestamp'] = utils.convert_to_argentina_time(lead['timestamp'])
    return jsonify({'success': True, 'leads': leads})


@profile_bp.route('/api/profile/lead/<int:lead_id>', methods=['GET'])
@decorators.login_required
def api_get_lead(lead_id):
    lang = get_language()
    user_id = session['user_id']
    lead = models.get_lead_by_id_and_user(lead_id, user_id)
    if not lead:
        return jsonify({'error': t('profile.lead_not_found', lang)}), 404
    if lead.get('timestamp'):
        lead['timestamp'] = utils.convert_to_argentina_time(lead['timestamp'])
    return jsonify({'success': True, 'lead': lead})


@profile_bp.route('/api/profile/lead/<int:lead_id>', methods=['PUT'])
@decorators.login_required
@rate_limit.check_rate_limit(limit=100, window=60)
def api_update_lead(lead_id):
    lang = get_language()
    user_id = session['user_id']
    data = request.json or {}

    lead = models.get_lead_by_id_and_user(lead_id, user_id)
    if not lead:
        return jsonify({'error': t('profile.lead_not_found', lang)}), 404

    allowed = ALLOWED_LEAD_EDIT_FIELDS
    update_data = {k: utils.safe_text(v) for k, v in data.items() if k in allowed}

    if not update_data:
        return jsonify({'error': t('profile.no_valid_data', lang)}), 400

    snapshot = json.dumps({k: str(lead.get(k, '')) for k in allowed})
    max_ver = models.get_lead_max_version(lead_id)
    models.create_lead_version(lead_id, max_ver + 1, snapshot, user_id, '')

    models.update_lead(lead_id, update_data)

    utils.log_action('Edicion de Lead', f'Lead ID: {lead_id} editado por {session["username"]}', session)

    return jsonify({'status': 'success', 'message': t('profile.lead_updated', lang)})


@profile_bp.route('/api/profile/lead/<int:lead_id>/versions', methods=['GET'])
@decorators.login_required
def api_get_lead_versions(lead_id):
    lang = get_language()
    user_id = session['user_id']
    lead = models.get_lead_by_id_and_user(lead_id, user_id)
    if not lead:
        return jsonify({'error': t('profile.lead_not_found', lang)}), 404
    versions = models.get_lead_versions(lead_id)
    for v in versions:
        if v.get('edited_at'):
            v['edited_at'] = utils.convert_to_argentina_time(v['edited_at'])
    return jsonify({'success': True, 'versions': versions})


@profile_bp.route('/api/profile/user', methods=['GET'])
@decorators.login_required
def api_get_user():
    lang = get_language()
    user = models.get_user_profile(session['user_id'])
    if not user:
        return jsonify({'error': t('profile.user_not_found', lang)}), 404
    return jsonify({'success': True, 'user': user})


@profile_bp.route('/api/profile/user', methods=['PUT'])
@decorators.login_required
@rate_limit.check_rate_limit(limit=100, window=60)
def api_update_user():
    lang = get_language()
    user_id = session['user_id']
    data = request.json or {}

    email = utils.safe_text(data.get('email', '')).strip()
    phone = (data.get('phone', '') or '').strip()
    first_name = utils.safe_text(data.get('first_name', '')).strip()
    last_name = utils.safe_text(data.get('last_name', '')).strip()
    bio = utils.safe_text(data.get('bio', '')).strip()

    if email:
        is_valid, error = validators.validate_email(email)
        if not is_valid:
            return jsonify({'error': error}), 400

    if phone:
        is_valid, error = validators.validate_phone(phone)
        if not is_valid:
            return jsonify({'error': error}), 400

    models.update_user_credentials(user_id, email, phone)
    models.update_user_profile(user_id, {
        'first_name': first_name,
        'last_name': last_name,
        'bio': bio,
    })

    session['email'] = email
    session['phone'] = phone

    utils.log_action('Actualizacion de Perfil', f'Usuario: {session["username"]}', session)

    return jsonify({'status': 'success', 'message': t('profile.updated', lang)})


@profile_bp.route('/api/profile/user/password', methods=['PUT'])
@decorators.login_required
@rate_limit.check_rate_limit(limit=100, window=60)
def api_change_password():
    lang = get_language()
    user_id = session['user_id']
    data = request.json or {}

    current_password = data.get('current_password', '')
    new_password = data.get('new_password', '')

    if not current_password or not new_password:
        return jsonify({'error': t('profile.all_fields_required', lang)}), 400

    user = models.get_user_by_id(user_id)
    if not user or not check_password_hash(user['hash'], current_password):
        return jsonify({'error': t('profile.wrong_password', lang)}), 400

    is_valid, error = validators.validate_password(new_password)
    if not is_valid:
        return jsonify({'error': error}), 400

    conn = models.get_db_connection()
    try:
        conn.execute('UPDATE users SET hash = ? WHERE id = ?',
                     (generate_password_hash(new_password), user_id))
        conn.commit()
    finally:
        conn.close()

    utils.log_action('Cambio de Contrasena', f'Usuario: {session["username"]}', session)

    return jsonify({'status': 'success', 'message': t('profile.password_updated', lang)})


@profile_bp.route('/api/profile/professional', methods=['GET'])
@decorators.professional_required
def api_get_professional():
    lang = get_language()
    user_id = session['user_id']
    pro = models.get_professional_by_user_id(user_id)
    if not pro:
        return jsonify({'error': t('profile.pro_not_found', lang)}), 404
    return jsonify({'success': True, 'professional': pro})


@profile_bp.route('/api/profile/professional', methods=['PUT'])
@decorators.professional_required
@rate_limit.check_rate_limit(limit=100, window=60)
def api_update_professional():
    lang = get_language()
    user_id = session['user_id']
    data = request.json or {}

    ALLOWED_FIELDS = {'specialty', 'title', 'province', 'zone', 'country'}
    update_data = {}
    for field in ALLOWED_FIELDS:
        if field in data:
            update_data[field] = utils.safe_text(data[field]).strip()

    if not update_data:
        return jsonify({'error': t('profile.no_valid_data', lang)}), 400

    models.update_professional_profile(user_id, update_data)

    utils.log_action('Actualizacion Profesional', f'Usuario: {session["username"]}', session)

    return jsonify({'status': 'success', 'message': t('profile.pro_updated', lang)})


# ============================================================
# NUEVOS ENDPOINTS — Preferencias, Sesiones, Actividad
# ============================================================


@profile_bp.route('/api/profile/settings', methods=['GET'])
@decorators.login_required
def api_get_settings():
    prefs = models.get_user_preferences(session['user_id'])
    return jsonify({'success': True, 'preferences': prefs})


@profile_bp.route('/api/profile/settings', methods=['PUT'])
@decorators.login_required
@rate_limit.check_rate_limit(limit=100, window=60)
def api_update_settings():
    lang = get_language()
    user_id = session['user_id']
    data = request.json or {}

    allowed = {'theme', 'language', 'email_notifications', 'sms_notifications', 'lead_alerts', 'preferred_channel'}
    update_data = {}

    for key in allowed:
        if key in data:
            update_data[key] = data[key]

    if not update_data:
        return jsonify({'error': t('profile.no_valid_data', lang)}), 400

    if 'theme' in update_data and update_data['theme'] not in ('light', 'dark'):
        return jsonify({'error': t('profile.invalid_theme', lang)}), 400

    if 'language' in update_data and update_data['language'] not in ('es', 'en'):
        return jsonify({'error': t('profile.invalid_language', lang)}), 400

    if 'preferred_channel' in update_data and update_data['preferred_channel'] not in ('sms', 'whatsapp', 'auto'):
        return jsonify({'error': t('profile.invalid_channel', lang)}), 400

    models.update_user_preferences(user_id, update_data)
    utils.log_action('Actualizacion de Preferencias', f'Usuario: {session["username"]}', session)

    return jsonify({'status': 'success', 'message': t('profile.prefs_updated', lang)})


@profile_bp.route('/api/profile/sessions', methods=['GET'])
@decorators.login_required
def api_get_sessions():
    history = models.get_user_login_history(session['user_id'])
    return jsonify({'success': True, 'sessions': history})


@profile_bp.route('/api/profile/sessions/<int:entry_id>', methods=['DELETE'])
@decorators.login_required
@rate_limit.check_rate_limit(limit=100, window=60)
def api_delete_session(entry_id):
    lang = get_language()
    deleted = models.delete_login_history_entry(entry_id, session['user_id'])
    if not deleted:
        return jsonify({'error': t('profile.session_not_found', lang)}), 404
    utils.log_action('Sesion cerrada', f'Sesion ID: {entry_id}', session)
    return jsonify({'status': 'success', 'message': t('profile.session_closed', lang)})


@profile_bp.route('/api/profile/activity', methods=['GET'])
@decorators.login_required
def api_get_activity():
    limit = request.args.get('limit', 50, type=int)
    limit = min(limit, 100)
    activity = models.get_user_activity(session['user_id'], limit)
    for entry in activity:
        if entry.get('timestamp'):
            entry['timestamp'] = utils.convert_to_argentina_time(entry['timestamp'])
    return jsonify({'success': True, 'activity': activity})


@profile_bp.route('/api/profile/user/avatar', methods=['POST'])
@decorators.login_required
@rate_limit.check_rate_limit(limit=100, window=60)
def api_upload_avatar():
    lang = get_language()
    user_id = session['user_id']
    if 'avatar' not in request.files:
        return jsonify({'error': t('profile.no_file_sent', lang)}), 400

    file = request.files['avatar']
    if file.filename == '':
        return jsonify({'error': t('profile.empty_filename', lang)}), 400

    if not utils.allowed_file(file.filename):
        return jsonify({'error': t('profile.invalid_format', lang)}), 400

    mime_valid, detected_ext, mime_error = utils.validate_mime_type(file, file.filename)
    if not mime_valid:
        return jsonify({'error': mime_error}), 400

    file.seek(0, os.SEEK_END)
    size = file.tell()
    if size > config.MAX_UPLOAD_SIZE:
        return jsonify({'error': t('profile.file_too_large', lang)}), 400
    file.seek(0)

    ext = detected_ext or 'jpg'
    safe_filename = f'user_{user_id}_avatar.{ext}'
    avatar_dir = current_app.config['AVATAR_FOLDER']
    filepath = os.path.join(avatar_dir, safe_filename)

    # Eliminar avatar anterior si existe
    old_path = models.get_user_avatar_path(user_id)
    if old_path:
        old_file = os.path.join(avatar_dir, os.path.basename(old_path))
        if os.path.exists(old_file):
            os.remove(old_file)

    file.save(filepath)

    avatar_rel = f'uploads/avatars/{safe_filename}'
    models.update_user_avatar(user_id, avatar_rel)
    utils.log_action('Avatar actualizado', f'Usuario: {session["username"]}', session)

    return jsonify({'status': 'success', 'avatar_url': url_for('static', filename=avatar_rel)})


@profile_bp.route('/api/profile/user/avatar', methods=['DELETE'])
@decorators.login_required
@rate_limit.check_rate_limit(limit=100, window=60)
def api_delete_avatar():
    lang = get_language()
    user_id = session['user_id']
    old_path = models.get_user_avatar_path(user_id)
    if old_path:
        old_file = os.path.join(current_app.config['AVATAR_FOLDER'], os.path.basename(old_path))
        if os.path.exists(old_file):
            os.remove(old_file)
    models.delete_user_avatar(user_id)
    utils.log_action('Avatar eliminado', f'Usuario: {session["username"]}', session)
    return jsonify({'status': 'success', 'message': t('profile.avatar_deleted', lang)})


# ============================================================
# NUEVOS ENDPOINTS — Perfil Profesional Extendido
# ============================================================


@profile_bp.route('/api/profile/professional/full', methods=['GET'])
@decorators.professional_required
def api_get_professional_full():
    lang = get_language()
    user_id = session['user_id']
    pro = models.get_professional_full_profile(user_id)
    if not pro:
        return jsonify({'error': t('profile.pro_not_found', lang)}), 404
    return jsonify({'success': True, 'professional': pro})


@profile_bp.route('/api/profile/professional/full', methods=['PUT'])
@decorators.professional_required
@rate_limit.check_rate_limit(limit=100, window=60)
def api_update_professional_full():
    lang = get_language()
    user_id = session['user_id']
    data = request.json or {}

    allowed = {
        'bio_pro', 'experience_years', 'services_offered',
        'portfolio', 'availability', 'social_links',
        'fee_range_min', 'fee_range_max', 'professional_address'
    }
    update_data = {}
    for key in allowed:
        if key in data:
            val = data[key]
            if isinstance(val, str):
                val = utils.safe_text(val).strip()
            update_data[key] = val

    if not update_data:
        return jsonify({'error': t('profile.no_valid_data', lang)}), 400
    models.create_or_update_professional_profile(user_id, update_data)

    geo_fields = {}
    for key in ('province', 'zone', 'country'):
        if key in data:
            geo_fields[key] = utils.safe_text(data[key]).strip()
    if geo_fields:
        models.update_professional_profile(user_id, geo_fields)

    utils.log_action('Perfil profesional actualizado', f'Usuario: {session["username"]}', session)

    return jsonify({'status': 'success', 'message': t('profile.pro_updated', lang)})


@profile_bp.route('/api/profile/professional/photo', methods=['POST'])
@decorators.professional_required
@rate_limit.check_rate_limit(limit=100, window=60)
def api_upload_professional_photo():
    lang = get_language()
    user_id = session['user_id']
    if 'photo' not in request.files:
        return jsonify({'error': t('profile.no_file_sent', lang)}), 400

    file = request.files['photo']
    if file.filename == '':
        return jsonify({'error': t('profile.empty_filename', lang)}), 400

    if not utils.allowed_file(file.filename):
        return jsonify({'error': t('profile.invalid_format', lang)}), 400

    mime_valid, detected_ext, mime_error = utils.validate_mime_type(file, file.filename)
    if not mime_valid:
        return jsonify({'error': mime_error}), 400

    file.seek(0, os.SEEK_END)
    size = file.tell()
    if size > config.MAX_UPLOAD_SIZE:
        return jsonify({'error': t('profile.file_too_large', lang)}), 400
    file.seek(0)

    ext = detected_ext or 'jpg'
    safe_filename = f'pro_{user_id}_photo.{ext}'
    avatar_dir = current_app.config['AVATAR_FOLDER']
    filepath = os.path.join(avatar_dir, safe_filename)

    old_path = models.get_professional_photo_path(user_id)
    if old_path:
        old_file = os.path.join(avatar_dir, os.path.basename(old_path))
        if os.path.exists(old_file):
            os.remove(old_file)

    file.save(filepath)

    photo_rel = f'uploads/avatars/{safe_filename}'
    models.create_or_update_professional_profile(user_id, {'photo_path': photo_rel})
    utils.log_action('Foto profesional actualizada', f'Usuario: {session["username"]}', session)

    return jsonify({'status': 'success', 'photo_url': url_for('static', filename=photo_rel)})


@profile_bp.route('/api/profile/professional/photo', methods=['DELETE'])
@decorators.professional_required
@rate_limit.check_rate_limit(limit=100, window=60)
def api_delete_professional_photo():
    lang = get_language()
    user_id = session['user_id']
    old_path = models.get_professional_photo_path(user_id)
    if old_path:
        old_file = os.path.join(current_app.config['AVATAR_FOLDER'], os.path.basename(old_path))
        if os.path.exists(old_file):
            os.remove(old_file)
    models.create_or_update_professional_profile(user_id, {'photo_path': ''})
    utils.log_action('Foto profesional eliminada', f'Usuario: {session["username"]}', session)
    return jsonify({'status': 'success', 'message': t('profile.photo_deleted', lang)})


# ============================================================
# NOTIFICACIONES
# ============================================================


@profile_bp.route('/api/profile/notifications')
@decorators.login_required
def profile_notifications():
    user_id = session['user_id']
    notifications = models.get_user_notifications(user_id)
    unread = models.get_unread_notification_count(user_id)
    return jsonify({
        'success': True,
        'notifications': notifications,
        'unread': unread,
        'current_user_id': user_id,
    })


@profile_bp.route('/api/profile/notifications/read', methods=['POST'])
@decorators.login_required
def mark_notification_read():
    user_id = session['user_id']
    data = request.json or {}
    nid = data.get('notification_id')
    if not nid:
        return jsonify({'success': False, 'error': 'notification_id requerido'}), 400
    ok = models.mark_notification_read(nid, user_id)
    return jsonify({'success': ok})


@profile_bp.route('/api/profile/notifications/read-all', methods=['POST'])
@decorators.login_required
def mark_all_notifications_read():
    user_id = session['user_id']
    models.mark_all_notifications_read(user_id)
    return jsonify({'success': True})


@profile_bp.route('/api/profile/notifications/all')
@decorators.login_required
def all_notifications():
    user_id = session['user_id']
    page = request.args.get('page', 1, type=int)
    result = models.get_all_user_notifications(user_id, page=page)
    unread = models.get_unread_notification_count(user_id)
    return jsonify({
        'success': True,
        'notifications': result['items'],
        'total': result['total'],
        'page': result['page'],
        'pages': result['pages'],
        'unread': unread,
        'current_user_id': user_id,
    })


@profile_bp.route('/api/profile/notifications/delete-read', methods=['POST'])
@decorators.login_required
def delete_read_notifications():
    user_id = session['user_id']
    count = models.delete_read_notifications(user_id)
    lang = get_language()
    return jsonify({
        'success': True,
        'deleted': count,
        'message': t('notif.read_deleted', lang),
    })


@profile_bp.route('/api/profile/notifications/<int:notification_id>', methods=['DELETE'])
@decorators.login_required
def delete_notification(notification_id):
    user_id = session['user_id']
    ok = models.delete_notification(notification_id, user_id)
    return jsonify({'success': ok})


@profile_bp.route('/api/profile/notification-filters', methods=['PUT'])
@decorators.professional_required
def update_notification_filters():
    lang = get_language()
    user_id = session['user_id']
    data = request.json or {}
    filters = {
        'types': data.get('types', []),
        'property_types': data.get('property_types', []),
    }
    # Validate types against form_options
    valid_types = set(models.get_form_options_by_category('operation_type'))
    valid_prop_types = set(models.get_form_options_by_category('property_type'))
    filters['types'] = [ft for ft in filters['types'] if ft in valid_types]
    filters['property_types'] = [pt for pt in filters['property_types'] if pt in valid_prop_types]
    # Budget range
    budget_min = data.get('budget_min')
    budget_max = data.get('budget_max')
    updates = {'notification_filters': json.dumps(filters, ensure_ascii=False)}
    if budget_min is not None:
        try:
            updates['budget_min'] = max(0, float(budget_min))
        except (ValueError, TypeError):
            return jsonify({'success': False, 'error': t('profile.invalid_budget_range', lang)}), 400
    if budget_max is not None:
        try:
            updates['budget_max'] = max(0, float(budget_max))
        except (ValueError, TypeError):
            return jsonify({'success': False, 'error': t('profile.invalid_budget_range', lang)}), 400
    ok = models.update_user_preferences(user_id, updates)
    if ok:
        return jsonify({'success': True, 'filters': filters, 'budget_min': updates.get('budget_min', 0), 'budget_max': updates.get('budget_max', 0)})
    return jsonify({'success': False, 'error': t('profile.save_error', lang)}), 500


@profile_bp.route('/api/profile/notification-filters', methods=['GET'])
@decorators.login_required
def get_notification_filters():
    user_id = session['user_id']
    prefs = models.get_user_preferences(user_id)
    raw = prefs.get('notification_filters', '')
    filters = {}
    if raw:
        try:
            filters = json.loads(raw) if isinstance(raw, str) else raw
        except (json.JSONDecodeError, TypeError):
            filters = {}
    return jsonify({
        'success': True,
        'filters': filters,
        'budget_min': prefs.get('budget_min', 0),
        'budget_max': prefs.get('budget_max', 0),
    })


@profile_bp.route('/api/profile/notification-channel', methods=['PUT'])
@decorators.login_required
def update_notification_channel():
    lang = get_language()
    user_id = session['user_id']
    data = request.json or {}
    channel = data.get('channel', '').strip().lower()
    valid_channels = {'email', 'whatsapp', 'ambos', 'auto'}
    if channel not in valid_channels:
        return jsonify({'success': False, 'error': t('profile.invalid_channel', lang)}), 400
    ok = models.update_user_preferences(user_id, {'preferred_channel': channel})
    if ok:
        return jsonify({'success': True, 'channel': channel})
    return jsonify({'success': False, 'error': t('profile.save_error', lang)}), 500


@profile_bp.route('/api/profile/notification-channel', methods=['GET'])
@decorators.login_required
def get_notification_channel():
    user_id = session['user_id']
    prefs = models.get_user_preferences(user_id)
    return jsonify({
        'success': True,
        'channel': prefs.get('preferred_channel', 'email'),
        'whatsapp_notifications': prefs.get('whatsapp_notifications', 1),
    })


@profile_bp.route('/api/profile/contact-admin', methods=['POST'])
@decorators.login_required
def contact_admin():
    lang = get_language()
    data = request.json or {}
    subject = (data.get('subject') or '').strip()
    message = (data.get('message') or '').strip()

    if not subject or not message:
        return jsonify({"error": t('profile.contact_admin_required', lang)}), 400

    user = models.get_user_by_id(session['user_id'])
    username = user['username'] if user else 'Desconocido'

    title = t('profile.admin_msg_title', lang, user=username, subject=subject)
    body = t('profile.admin_msg_body', lang, message=message)

    notified = notify_admins(title=title, body=body)
    utils.log_action('Contacto admin desde perfil', f'{username}: {subject}', session)
    return jsonify({"status": "success", "message": t('profile.contact_admin_sent', lang)})


# ============================================================
# EXPORTACIÓN DE REPORTES
# ============================================================


def _query_monthly_contacts(user_id):
    """Retorna leads contactados en los últimos 30 días para un profesional."""
    conn = None
    try:
        conn = models.get_db_connection()
        rows = conn.execute('''
            SELECT l.id, l.type, l.property_type, l.zone, l.province,
                   l.budget, l.currency, l.timestamp,
                   lt.seen, lt.contacted, lt.seen_at, lt.contacted_at
            FROM leads l
            JOIN lead_tracking lt ON l.id = lt.lead_id
            WHERE lt.professional_id = ?
              AND lt.seen = 1
              AND lt.contacted = 1
              AND lt.contacted_at >= datetime('now', '-30 days')
            ORDER BY lt.contacted_at DESC
            LIMIT 1000
        ''', (user_id,)).fetchall()
        return [dict(r) for r in rows]
    finally:
        if conn:
            conn.close()



def _build_export_stats(leads):
    """Construye stats a partir de la lista de leads."""
    total = len(leads)
    budgets = []
    zones = {}
    currencies = {}
    types = {}
    budget_by_currency = {}
    for lead in leads:
        b = parse_budget(lead.get('budget'))
        budgets.append(b)
        z = lead.get('zone') or 'Sin zona'
        zones[z] = zones.get(z, 0) + 1
        c = lead.get('currency') or 'ARG'
        currencies[c] = currencies.get(c, 0) + 1
        t = lead.get('type') or 'Otro'
        types[t] = types.get(t, 0) + 1
        by_curr = budget_by_currency.setdefault(c, {'count': 0, 'total': 0.0})
        by_curr['count'] += 1
        by_curr['total'] += b
    avg_budget = round(sum(budgets) / total, 2) if total else 0
    total_budget = sum(budgets)
    sorted_zones = sorted(zones.items(), key=lambda x: x[1], reverse=True)
    return {
        'total': total,
        'avg_budget': avg_budget,
        'total_budget': total_budget,
        'zones': sorted_zones,
        'currencies': currencies,
        'types': types,
        'zone_count': len(zones),
        'budget_by_currency': budget_by_currency,
    }


@profile_bp.route('/api/profile/professional/export/xlsx')
@decorators.professional_required
def profile_export_history_xlsx():
    try:
        lang = get_language()
        user_id = session['user_id']
        leads = _query_monthly_contacts(user_id)
        stats = _build_export_stats(leads)

        wb = Workbook()
        data_font = Font(name='Manrope', size=10, color='000410')
        label_font = Font(name='Manrope', size=10, bold=True, color='735A3A')

        # ─── Sheet 1: Resumen ───
        ws_resumen = wb.active
        ws_resumen.title = t('profile.xlsx_summary', lang)
        resumen_data = [
            [t('profile.xlsx_total_leads_contacted', lang), stats['total']],
            [t('profile.xlsx_avg_budget', lang), f"${stats['avg_budget']:,.2f}"],
            [t('profile.xlsx_total_budget', lang), f"${stats['total_budget']:,.2f}"],
            [t('profile.xlsx_different_zones', lang), stats['zone_count']],
        ]
        for col, h in enumerate([t('profile.xlsx_metric', lang), t('profile.xlsx_value', lang)], 1):
            ws_resumen.cell(row=1, column=col, value=h)
        _style_header_row(ws_resumen, 2)
        ws_resumen.column_dimensions['A'].width = 32
        ws_resumen.column_dimensions['B'].width = 22
        for row_idx, row_data in enumerate(resumen_data, 2):
            ws_resumen.cell(row=row_idx, column=1, value=row_data[0]).font = label_font
            ws_resumen.cell(row=row_idx, column=2, value=row_data[1]).font = data_font
            ws_resumen.cell(row=row_idx, column=2).alignment = Alignment(horizontal='center')
            _apply_data_border(ws_resumen, row_idx, 2)

        note_row = len(resumen_data) + 3
        ws_resumen.cell(row=note_row, column=1,
            value=t('profile.xlsx_generated_by', lang)
        ).font = Font(name='Manrope', size=8, italic=True, color='A68A64')

        # ─── Sheet 2: Leads ───
        ws_leads = wb.create_sheet('Leads')
        headers = ['ID', t('profile.xlsx_type', lang), t('profile.xlsx_property', lang), t('profile.xlsx_zone', lang), t('profile.xlsx_province', lang), t('profile.xlsx_budget', lang),
                   t('profile.xlsx_currency', lang), t('profile.xlsx_created', lang), t('profile.xlsx_seen', lang), t('profile.xlsx_contacted', lang)]
        for col, h in enumerate(headers, 1):
            ws_leads.cell(row=1, column=col, value=h)
        _style_header_row(ws_leads, len(headers))
        col_widths = [8, 22, 16, 20, 18, 14, 10, 18, 18, 18]
        for i, w in enumerate(col_widths, 1):
            ws_leads.column_dimensions[get_column_letter(i)].width = w
        ws_leads.auto_filter.ref = f'A1:J{len(leads) + 1}'
        for row_idx, lead in enumerate(leads, 2):
            vals = [
                lead['id'], lead['type'], lead['property_type'], lead['zone'],
                lead.get('province', ''), lead['budget'], lead['currency'],
                lead.get('timestamp', ''), lead.get('seen_at', ''), lead.get('contacted_at', ''),
            ]
            for col, val in enumerate(vals, 1):
                cell = ws_leads.cell(row=row_idx, column=col, value=val)
                cell.font = data_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            _apply_data_border(ws_leads, row_idx, len(headers))

        # ─── Sheet 3: Por Zona ───
        ws_zones = wb.create_sheet('Por Zona')
        for col, h in enumerate([t('profile.xlsx_zone', lang), t('profile.xlsx_count', lang), t('profile.xlsx_pct_total', lang)], 1):
            ws_zones.cell(row=1, column=col, value=h)
        _style_header_row(ws_zones, 3)
        ws_zones.column_dimensions['A'].width = 30
        ws_zones.column_dimensions['B'].width = 14
        ws_zones.column_dimensions['C'].width = 14
        ws_zones.auto_filter.ref = f'A1:C{len(stats["zones"]) + 1}'
        for row_idx, (zone, count) in enumerate(stats['zones'], 2):
            pct = round(count / stats['total'] * 100, 1) if stats['total'] else 0
            ws_zones.cell(row=row_idx, column=1, value=zone).font = label_font
            ws_zones.cell(row=row_idx, column=2, value=count).font = data_font
            ws_zones.cell(row=row_idx, column=2).alignment = Alignment(horizontal='center')
            ws_zones.cell(row=row_idx, column=3, value=f'{pct}%').font = data_font
            ws_zones.cell(row=row_idx, column=3).alignment = Alignment(horizontal='center')
            _apply_data_border(ws_zones, row_idx, 3)

        # ─── Sheet 4: Inversiones ───
        ws_inv = wb.create_sheet('Inversiones')
        for col, h in enumerate([t('profile.xlsx_currency', lang), t('profile.xlsx_count', lang), t('profile.xlsx_total_budget_col', lang), t('profile.xlsx_avg_budget_col', lang)], 1):
            ws_inv.cell(row=1, column=col, value=h)
        _style_header_row(ws_inv, 4)
        ws_inv.column_dimensions['A'].width = 16
        ws_inv.column_dimensions['B'].width = 14
        ws_inv.column_dimensions['C'].width = 22
        ws_inv.column_dimensions['D'].width = 24
        ws_inv.auto_filter.ref = f'A1:D{len(stats["currencies"]) + 1}'
        row_idx = 2
        for currency, count in sorted(stats['currencies'].items()):
            bc = stats['budget_by_currency'].get(currency, {'count': 0, 'total': 0.0})
            total_c = bc['total']
            avg_c = round(total_c / bc['count'], 2) if bc['count'] else 0
            ws_inv.cell(row=row_idx, column=1, value=currency).font = label_font
            ws_inv.cell(row=row_idx, column=2, value=count).font = data_font
            ws_inv.cell(row=row_idx, column=2).alignment = Alignment(horizontal='center')
            ws_inv.cell(row=row_idx, column=3, value=f"${total_c:,.2f}").font = data_font
            ws_inv.cell(row=row_idx, column=3).alignment = Alignment(horizontal='center')
            ws_inv.cell(row=row_idx, column=4, value=f"${avg_c:,.2f}").font = data_font
            ws_inv.cell(row=row_idx, column=4).alignment = Alignment(horizontal='center')
            _apply_data_border(ws_inv, row_idx, 4)
            row_idx += 1

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        now_str = datetime.now().strftime('%Y%m%d_%H%M')
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'reporte_rendimiento_{now_str}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception:
        logger.exception('Error exporting XLSX report')
        lang = get_language()
        return jsonify({'error': t('profile.xlsx_export_error', lang)}), 500


@profile_bp.route('/api/profile/professional/export/pdf')
@decorators.professional_required
def profile_export_history_pdf():
    try:
        lang = get_language()
        user_id = session['user_id']
        leads = _query_monthly_contacts(user_id)
        stats = _build_export_stats(leads)

        # Fetch professional name
        conn = models.get_db_connection()
        pro = conn.execute(
            'SELECT name FROM professionals WHERE user_id = ?', (user_id,)
        ).fetchone()
        prof_name = pdf_safe(pro['name']) if pro else 'Profesional'
        conn.close()

        midnight = (0, 4, 16)
        gold = (115, 90, 58)
        gold_light = (166, 138, 100)
        gray_bg = (245, 243, 240)
        white = (255, 255, 255)

        class ReportPDF(FPDF):
            def header(self):
                if self.page_no() == 1:
                    return
                self.set_fill_color(*midnight)
                self.rect(0, 0, 210, 10, 'F')
                self.set_font('Helvetica', 'B', 7)
                self.set_text_color(*gold_light)
                self.set_y(2)
                self.cell(0, 6, t('profile.pdf_report_title', lang), ln=True, align='C')
                self.ln(4)

            def footer(self):
                self.set_y(-12)
                self.set_font('Helvetica', 'I', 7)
                self.set_text_color(150, 150, 150)
                self.cell(0, 8, f'{t("profile.pdf_page", lang)} {self.page_no()}/{{nb}}  |  {datetime.now().strftime("%d/%m/%Y %H:%M")}', ln=True, align='C')

        pdf = ReportPDF()
        pdf.alias_nb_pages()
        pdf.set_auto_page_break(auto=True, margin=18)

        # ─── PAGE 1 — Brand Header ───
        pdf.add_page()
        pdf.set_fill_color(*midnight)
        pdf.rect(0, 0, 210, 28, 'F')

        pdf.set_y(4)
        pdf.set_font('Times', 'BI', 18)
        pdf.set_text_color(*gold_light)
        pdf.cell(0, 10, 'ArchEstate', ln=True, align='C')
        pdf.set_font('Helvetica', '', 7)
        pdf.set_text_color(166, 138, 100)
        pdf.cell(0, 5, 'The Private Ledger', ln=True, align='C')

        pdf.set_y(24)
        pdf.set_draw_color(*gold)
        pdf.set_line_width(0.5)
        pdf.line(20, 28, 190, 28)

        # ─── Professional info ───
        pdf.set_y(32)
        pdf.set_font('Helvetica', 'B', 10)
        pdf.set_text_color(*midnight)
        pdf.cell(0, 6, pdf_safe(prof_name), ln=True, align='L')
        pdf.set_font('Helvetica', '', 8)
        pdf.set_text_color(100, 100, 100)
        pdf.cell(0, 5, t('profile.pdf_subtitle', lang), ln=True, align='L')
        pdf.ln(6)

        # ─── Summary cards (2x2) ───
        card_w = 82
        card_h = 18
        gap = 6
        left_margin = 14
        pdf.set_font('Helvetica', '', 9)

        cards = [
            (t('profile.pdf_total_leads', lang), str(stats['total']), f"{stats['zone_count']} {t('profile.pdf_active_zones', lang)}"),
            (t('profile.pdf_total_budget', lang), f"${stats['total_budget']:,.2f}", f"{t('profile.pdf_average', lang)} ${stats['avg_budget']:,.2f}"),
        ]
        pdf.set_x(left_margin)
        for title, value, sub in cards:
            x0 = pdf.get_x()
            y0 = pdf.get_y()
            pdf.set_fill_color(*gold)
            pdf.rect(x0, y0, 4, card_h, 'F')
            pdf.set_fill_color(*gray_bg)
            pdf.rect(x0 + 4, y0, card_w - 4, card_h, 'F')
            pdf.set_text_color(*midnight)
            pdf.set_xy(x0 + 8, y0 + 2)
            pdf.set_font('Helvetica', '', 7)
            pdf.cell(card_w - 12, 4, pdf_safe(title))
            pdf.set_xy(x0 + 8, y0 + 6)
            pdf.set_font('Helvetica', 'B', 13)
            pdf.cell(card_w - 12, 7, pdf_safe(value))
            pdf.set_xy(x0 + 8, y0 + 13)
            pdf.set_font('Helvetica', 'I', 7)
            pdf.set_text_color(100, 100, 100)
            pdf.cell(card_w - 12, 4, pdf_safe(sub))
            pdf.set_x(x0 + card_w + gap)

        pdf.ln(card_h + 4)
        pdf.set_text_color(*midnight)

        pdf.set_x(left_margin)
        cards2 = [
            (t('profile.pdf_avg_per_lead', lang), f"${stats['avg_budget']:,.2f}", t('profile.pdf_estimated_avg', lang)),
            (t('profile.pdf_active_currencies', lang), str(len(stats['currencies'])), f"{', '.join(sorted(stats['currencies'].keys()))}"),
        ]
        for title, value, sub in cards2:
            x0 = pdf.get_x()
            y0 = pdf.get_y()
            pdf.set_fill_color(*gold)
            pdf.rect(x0, y0, 4, card_h, 'F')
            pdf.set_fill_color(*gray_bg)
            pdf.rect(x0 + 4, y0, card_w - 4, card_h, 'F')
            pdf.set_text_color(*midnight)
            pdf.set_xy(x0 + 8, y0 + 2)
            pdf.set_font('Helvetica', '', 7)
            pdf.cell(card_w - 12, 4, pdf_safe(title))
            pdf.set_xy(x0 + 8, y0 + 6)
            pdf.set_font('Helvetica', 'B', 13)
            pdf.cell(card_w - 12, 7, pdf_safe(value))
            pdf.set_xy(x0 + 8, y0 + 13)
            pdf.set_font('Helvetica', 'I', 7)
            pdf.set_text_color(100, 100, 100)
            pdf.cell(card_w - 12, 4, pdf_safe(sub))
            pdf.set_x(x0 + card_w + gap)

        pdf.ln(card_h + 8)

        # ─── Currency breakdown ───
        def section_header(title):
            pdf.set_fill_color(*gold)
            pdf.set_text_color(*white)
            pdf.set_font('Helvetica', 'B', 10)
            pdf.cell(0, 7, pdf_safe('  ' + title.upper()), ln=True, fill=True)
            pdf.set_text_color(*midnight)
            pdf.set_font('Helvetica', '', 9)
            pdf.ln(2)

        section_header(t('profile.pdf_by_currency', lang))
        pdf.set_font('Helvetica', 'B', 8)
        pdf.set_fill_color(*gray_bg)
        cw = [30, 30, 50, 50, 30]
        ch = [t('profile.pdf_currency', lang), t('profile.pdf_count', lang), t('profile.pdf_total_budget_label', lang), t('profile.pdf_avg_budget_label', lang), t('profile.pdf_pct', lang)]
        for i, (h, w) in enumerate(zip(ch, cw)):
            pdf.cell(w, 6, pdf_safe(h), border=1, align='C', fill=True)
        pdf.ln()
        pdf.set_font('Helvetica', '', 8)
        for currency, count in sorted(stats['currencies'].items()):
            bc = stats['budget_by_currency'].get(currency, {'count': 0, 'total': 0.0})
            total_c = bc['total']
            avg_c = round(total_c / bc['count'], 2) if bc['count'] else 0
            pct = round(count / stats['total'] * 100, 1) if stats['total'] else 0
            vals = [currency, str(count), f"${total_c:,.2f}", f"${avg_c:,.2f}", f"{pct}%"]
            for v, w in zip(vals, cw):
                pdf.cell(w, 5, pdf_safe(v), border=1, align='C')
            pdf.ln()
        pdf.ln(5)

        # ─── Property type breakdown ───
        section_header(t('profile.pdf_by_property_type', lang))
        pdf.set_font('Helvetica', 'B', 8)
        pdf.set_fill_color(*gray_bg)
        cw2 = [80, 40, 40, 30]
        ch2 = [t('profile.pdf_type', lang), t('profile.pdf_count', lang), t('profile.pdf_pct_total', lang), t('profile.pdf_avg_budget', lang)]
        for h, w in zip(ch2, cw2):
            pdf.cell(w, 6, pdf_safe(h), border=1, align='C', fill=True)
        pdf.ln()
        pdf.set_font('Helvetica', '', 8)

        # Compute per-type budgets
        type_budgets = {}
        for lead in leads:
            type_val = lead.get('type') or 'Otro'
            b = parse_budget(lead.get('budget'))
            tb = type_budgets.setdefault(type_val, {'count': 0, 'total': 0.0})
            tb['count'] += 1
            tb['total'] += b
        for type_val, count in sorted(stats['types'].items(), key=lambda x: x[1], reverse=True):
            tb = type_budgets.get(type_val, {'count': 0, 'total': 0.0})
            avg_t = round(tb['total'] / tb['count'], 2) if tb['count'] else 0
            pct = round(count / stats['total'] * 100, 1) if stats['total'] else 0
            vals = [pdf_safe(type_val)[:35], str(count), f"{pct}%", f"${avg_t:,.2f}"]
            for v, w in zip(vals, cw2):
                pdf.cell(w, 5, pdf_safe(v), border=1, align='C')
            pdf.ln()
        pdf.ln(5)

        # ─── Zone breakdown ───
        section_header(t('profile.pdf_top_zones', lang))
        pdf.set_font('Helvetica', 'B', 8)
        pdf.set_fill_color(*gray_bg)
        cw3 = [90, 40, 40, 22]
        ch3 = [t('profile.pdf_zone', lang), t('profile.pdf_count', lang), t('profile.pdf_pct_total', lang), t('profile.pdf_bar', lang)]
        for h, w in zip(ch3, cw3):
            pdf.cell(w, 6, pdf_safe(h), border=1, align='C', fill=True)
        pdf.ln()
        pdf.set_font('Helvetica', '', 8)
        max_count = max((c for _, c in stats['zones']), default=1)
        for zone, count in stats['zones'][:10]:
            pct = round(count / stats['total'] * 100, 1) if stats['total'] else 0
            vals = [pdf_safe(zone)[:40], str(count), f"{pct}%"]
            for v, w in zip(vals, cw3[:3]):
                pdf.cell(w, 5, pdf_safe(v), border=1, align='C')
            # Mini bar
            bar_ratio = count / max_count
            bar_w = int(bar_ratio * 18)
            x0 = pdf.get_x()
            y0 = pdf.get_y()
            pdf.set_fill_color(*gold_light)
            pdf.rect(x0, y0 + 1, bar_w, 3, 'F')
            pdf.cell(cw3[3], 5, '', border=1)
            pdf.ln()
        pdf.ln(5)

        # ─── Leads detail table ───
        section_header(t('profile.pdf_lead_detail', lang))
        pdf.set_font('Helvetica', 'B', 7)
        pdf.set_fill_color(*midnight)
        pdf.set_text_color(*white)
        cols_l = [8, 8, 32, 26, 24, 22, 28, 22]
        headers_l = ['#', 'ID', t('profile.pdf_type', lang), t('profile.pdf_zone', lang), t('profile.pdf_budget', lang), t('profile.pdf_currency', lang), t('profile.pdf_province', lang), t('profile.pdf_contacted_label', lang)]
        for h, w in zip(headers_l, cols_l):
            pdf.cell(w, 6, pdf_safe(h), border=1, align='C', fill=True)
        pdf.ln()
        pdf.set_text_color(*midnight)
        pdf.set_font('Helvetica', '', 7)
        for idx, lead in enumerate(leads[:30], 1):
            contact_date = lead.get('contacted_at', '')[:10] if lead.get('contacted_at') else ''
            budget = pdf_val(lead['budget'])
            vals = [
                str(idx), str(lead['id']), pdf_safe(lead['type'])[:28],
                pdf_safe(lead['zone'])[:22],
                budget[:20], pdf_safe(lead.get('currency', ''))[:8],
                pdf_safe(lead.get('province', ''))[:24],
                contact_date[:20],
            ]
            if idx % 2 == 0:
                pdf.set_fill_color(*gray_bg)
                fill = True
            else:
                fill = False
            for v, w in zip(vals, cols_l):
                pdf.cell(w, 5, pdf_safe(v), border=1, align='C', fill=fill)
            pdf.ln()
            if pdf.get_y() > 258:
                pdf.add_page()

        pdf.ln(5)
        pdf.set_font('Helvetica', 'I', 7)
        pdf.set_text_color(*gold_light)
        gen_date = datetime.now().strftime('%d/%m/%Y %H:%M')
        pdf.cell(0, 5, f'{t("profile.pdf_generated_by", lang)} | {gen_date}', ln=True, align='C')

        pdf_output = pdf.output(dest='S')
        if isinstance(pdf_output, str):
            pdf_output = pdf_output.encode('latin-1')

        buffer = BytesIO(pdf_output)
        buffer.seek(0)

        now_str = datetime.now().strftime('%Y%m%d_%H%M')
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'reporte_rendimiento_{now_str}.pdf',
            mimetype='application/pdf',
        )
    except Exception:
        logger.exception('Error exporting PDF report')
        lang = get_language()
        return jsonify({'error': t('profile.pdf_export_error', lang)}), 500
